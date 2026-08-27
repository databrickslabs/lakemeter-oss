"""Regression tests for VM pricing in Excel exports."""
from datetime import datetime
from io import BytesIO
from types import SimpleNamespace

import openpyxl
import pytest

from app.routes.export.excel_builder import (
    _lookup_dbsql_vm_costs,
    build_estimate_excel,
)
from tests.export.all_purpose.conftest import make_line_item as make_all_purpose_item
from tests.export.dbsql.conftest import make_line_item as make_dbsql_item
from tests.export.dlt.conftest import make_line_item as make_dlt_item
from tests.export.jobs.conftest import make_line_item as make_jobs_item


class _PricingDB:
    def __init__(self, rates):
        self.rates = rates
        self.calls = []

    def execute(self, statement, params):
        assert "sync_pricing_vm_costs" in str(statement)
        self.calls.append(params)
        key = (
            params["instance_type"].lower(),
            params["pricing_tier"],
            params["payment_option"].lower(),
        )
        rate = self.rates.get(key)
        row = (
            SimpleNamespace(cost_per_hour=rate, source="test pricing")
            if rate is not None
            else None
        )
        return SimpleNamespace(fetchone=lambda: row)


def _estimate():
    return SimpleNamespace(
        estimate_name="VM Pricing Regression",
        status="draft",
        version=1,
        created_at=datetime(2026, 8, 6),
        updated_at=datetime(2026, 8, 6),
    )


def _load_workbook(item, cloud, region, db=None):
    output = build_estimate_excel(
        _estimate(), [item], cloud, region, "PREMIUM", db=db
    )
    return openpyxl.load_workbook(BytesIO(output.read()), data_only=True)


def _find_row(sheet, name):
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value == name:
            return row
    raise AssertionError(f"Could not find workload row {name}")


def test_excel_uses_exact_reserved_and_spot_rates():
    db = _PricingDB({
        ("i3.xlarge", "reserved_1y", "all_upfront"): 0.229909,
        ("i3.xlarge", "spot", "na"): 0.123604,
    })
    item = make_jobs_item(
        workload_name="Exact VM Rates",
        driver_node_type="i3.xlarge",
        worker_node_type="i3.xlarge",
        num_workers=2,
        hours_per_month=100,
        driver_pricing_tier="reserved_1y",
        driver_payment_option="all_upfront",
        worker_pricing_tier="spot",
        worker_payment_option="NA",
    )

    workbook = _load_workbook(item, "aws", "us-east-1", db)
    sheet = workbook.active
    row = _find_row(sheet, "Exact VM Rates")

    assert sheet.cell(row=row, column=27).value == pytest.approx(0.229909)
    assert sheet.cell(row=row, column=28).value == pytest.approx(0.123604)
    assert "conservative fallback" not in (
        sheet.cell(row=row, column=34).value or ""
    )
    workbook.close()


@pytest.mark.parametrize(
    ("factory", "workload_name"),
    [
        (make_jobs_item, "Jobs VM Parity"),
        (make_all_purpose_item, "All-Purpose VM Parity"),
        (make_dlt_item, "DLT VM Parity"),
    ],
)
def test_all_classic_compute_exports_use_selected_vm_rates(factory, workload_name):
    db = _PricingDB({
        ("i3.xlarge", "reserved_1y", "no_upfront"): 0.229909,
        ("i3.xlarge", "spot", "na"): 0.123604,
    })
    item = factory(
        workload_name=workload_name,
        driver_node_type="i3.xlarge",
        worker_node_type="i3.xlarge",
        num_workers=2,
        hours_per_month=11,
        driver_pricing_tier="reserved_1y",
        driver_payment_option="no_upfront",
        worker_pricing_tier="spot",
        worker_payment_option="NA",
    )

    workbook = _load_workbook(item, "aws", "us-east-1", db)
    sheet = workbook.active
    row = _find_row(sheet, workload_name)

    assert sheet.cell(row=row, column=27).value == pytest.approx(0.229909)
    assert sheet.cell(row=row, column=28).value == pytest.approx(0.123604)
    assert sheet.cell(row=row, column=31).value == pytest.approx(5.248287)
    workbook.close()


def test_missing_azure_spot_rate_is_not_synthetically_discounted():
    item = make_jobs_item(
        workload_name="Azure Spot Fallback",
        driver_node_type="Standard_DS3_v2",
        worker_node_type="Standard_DS3_v2",
        num_workers=1,
        hours_per_month=100,
        driver_pricing_tier="spot",
        worker_pricing_tier="spot",
    )

    workbook = _load_workbook(item, "azure", "eastus")
    sheet = workbook.active
    row = _find_row(sheet, "Azure Spot Fallback")

    assert sheet.cell(row=row, column=27).value == pytest.approx(0.293)
    assert sheet.cell(row=row, column=28).value == pytest.approx(0.293)
    assert "conservative fallback" in sheet.cell(row=row, column=34).value
    workbook.close()


def test_dbsql_lookup_returns_rates_for_selected_payment_option():
    db = _PricingDB({
        ("i3.4xlarge", "reserved_3y", "no_upfront"): 0.31,
        ("i3.2xlarge", "reserved_3y", "no_upfront"): 0.16,
    })
    item = make_dbsql_item(
        dbsql_warehouse_type="CLASSIC",
        dbsql_warehouse_size="Small",
        dbsql_vm_pricing_tier="reserved_3y",
        dbsql_vm_payment_option="no_upfront",
    )
    notes = []

    driver_rate, worker_rate, worker_count = _lookup_dbsql_vm_costs(
        item, "aws", "us-east-1", {}, db, notes
    )

    assert driver_rate == pytest.approx(0.31)
    assert worker_rate == pytest.approx(0.16)
    assert worker_count == 4
    assert {call["payment_option"] for call in db.calls} == {"no_upfront"}
    assert notes == []


def test_dbsql_export_prefers_current_driver_worker_tiers_over_legacy_tier():
    db = _PricingDB({
        ("i3.4xlarge", "on_demand", "na"): 1.248,
        ("i3.2xlarge", "spot", "na"): 0.3072,
        ("i3.4xlarge", "reserved_1y", "all_upfront"): 0.75,
        ("i3.2xlarge", "reserved_1y", "all_upfront"): 0.20,
    })
    item = make_dbsql_item(
        workload_name="DBSQL UI Parity",
        dbsql_warehouse_type="CLASSIC",
        dbsql_warehouse_size="Small",
        hours_per_month=11,
        driver_pricing_tier="on_demand",
        driver_payment_option="no_upfront",
        worker_pricing_tier="spot",
        worker_payment_option="NA",
        # Simulate an estimate created before separate driver/worker controls.
        dbsql_vm_pricing_tier="reserved_1y",
        dbsql_vm_payment_option="all_upfront",
    )

    workbook = _load_workbook(item, "aws", "us-east-1", db)
    sheet = workbook.active
    row = _find_row(sheet, "DBSQL UI Parity")

    assert sheet.cell(row=row, column=10).value == "On-Demand"
    assert sheet.cell(row=row, column=11).value == "Spot"
    assert sheet.cell(row=row, column=27).value == pytest.approx(1.248)
    assert sheet.cell(row=row, column=28).value == pytest.approx(0.3072)
    assert sheet.cell(row=row, column=31).value == pytest.approx(27.2448)
    assert {call["pricing_tier"] for call in db.calls} == {"on_demand", "spot"}
    workbook.close()


def test_dbsql_export_defaults_reserved_worker_na_to_aws_no_upfront():
    db = _PricingDB({
        ("i3.4xlarge", "reserved_1y", "no_upfront"): 0.794977,
        ("i3.2xlarge", "reserved_1y", "no_upfront"): 0.405854,
    })
    item = make_dbsql_item(
        workload_name="DBSQL Reserved UI Parity",
        dbsql_warehouse_type="CLASSIC",
        dbsql_warehouse_size="Small",
        hours_per_month=11,
        driver_pricing_tier="reserved_1y",
        driver_payment_option="no_upfront",
        worker_pricing_tier="reserved_1y",
        # Reproduces the stale value saved while the UI displayed No Upfront.
        worker_payment_option="NA",
    )

    workbook = _load_workbook(item, "aws", "us-east-1", db)
    sheet = workbook.active
    row = _find_row(sheet, "DBSQL Reserved UI Parity")

    assert sheet.cell(row=row, column=10).value == "1-Year Reserved"
    assert sheet.cell(row=row, column=11).value == "1-Year Reserved"
    assert sheet.cell(row=row, column=27).value == pytest.approx(0.794977)
    assert sheet.cell(row=row, column=28).value == pytest.approx(0.405854)
    assert sheet.cell(row=row, column=31).value == pytest.approx(26.602323)
    assert {call["payment_option"] for call in db.calls} == {"no_upfront"}
    workbook.close()
