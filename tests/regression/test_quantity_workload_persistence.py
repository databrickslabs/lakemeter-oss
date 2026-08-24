from io import BytesIO
from datetime import datetime, timezone
from types import SimpleNamespace
from uuid import uuid4

import openpyxl
import pytest

from app.routes.export import build_estimate_excel
from app.routes.export.excel_item_helpers import calc_item_values
from app.schemas import LineItemCreate, LineItemResponse
from app.schemas.line_item import map_ai_parse_api_fields
from tests.regression.conftest import make_item
from tests.regression.excel_helpers import (
    COL_DBU_COST_D,
    COL_DBU_COST_L,
    COL_DBUS_MO,
    find_row_by_name,
    make_estimate,
)


def test_shutterstock_frontend_field_maps_to_storage_and_response():
    line_item = LineItemCreate(
        estimate_id=uuid4(),
        workload_name="Image generation",
        workload_type="SHUTTERSTOCK_IMAGEAI",
        shutterstock_images=500,
    )

    data = map_ai_parse_api_fields(
        line_item.model_dump(),
        line_item.model_fields_set,
    )

    assert data["shutterstock_imageai_num_images"] == 500
    assert "shutterstock_images" not in data

    response = LineItemResponse.model_validate({
        **data,
        "line_item_id": uuid4(),
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc),
    })
    assert response.shutterstock_images == 500


@pytest.mark.parametrize(
    ("item", "expected_dbus"),
    [
        (
            SimpleNamespace(
                workload_type="AI_PARSE",
                ai_parse_complexity="high",
                ai_parse_num_pages=10_000,
            ),
            875,
        ),
        (
            SimpleNamespace(
                workload_type="SHUTTERSTOCK_IMAGEAI",
                shutterstock_imageai_num_images=500,
            ),
            428.5,
        ),
    ],
)
def test_excel_uses_canonical_quantity_storage_fields(item, expected_dbus):
    _, _, _, total_dbus, _ = calc_item_values(
        item,
        is_fmapi_token=False,
        is_fmapi_provisioned=False,
        dbu_per_hour=0,
        cloud="aws",
        auto_notes=[],
    )

    assert total_dbus == pytest.approx(expected_dbus)


@pytest.mark.parametrize(
    ("item", "expected_dbus"),
    [
        (
            make_item(
                workload_type="AI_PARSE",
                workload_name="AI Parse pages",
                ai_parse_complexity="medium",
                ai_parse_num_pages=100_000,
            ),
            6_250,
        ),
        (
            make_item(
                workload_type="SHUTTERSTOCK_IMAGEAI",
                workload_name="Shutterstock images",
                shutterstock_imageai_num_images=50_000,
            ),
            42_850,
        ),
    ],
)
def test_quantity_exports_survive_formula_recalculation(item, expected_dbus):
    output = build_estimate_excel(
        make_estimate(),
        [item],
        cloud="aws",
        region="us-east-1",
        tier="PREMIUM",
    )
    workbook_bytes = output.getvalue()
    formulas = openpyxl.load_workbook(
        BytesIO(workbook_bytes),
        data_only=False,
    ).active
    cached = openpyxl.load_workbook(
        BytesIO(workbook_bytes),
        data_only=True,
    ).active

    row = find_row_by_name(formulas, item.workload_name)
    assert row is not None
    assert formulas.cell(row, COL_DBUS_MO).value == pytest.approx(expected_dbus)
    assert cached.cell(row, COL_DBUS_MO).value == pytest.approx(expected_dbus)
    assert formulas.cell(row, COL_DBU_COST_L).value == f"=Q{row}*R{row}"
    assert formulas.cell(row, COL_DBU_COST_D).value == f"=Q{row}*T{row}"


@pytest.mark.parametrize(
    ("item", "expected_columns"),
    [
        (
            make_item(
                workload_type="JOBS",
                workload_name="Hourly workload",
                serverless_enabled=True,
                serverless_mode="standard",
                hours_per_month=100,
            ),
            ("P", "L"),
        ),
        (
            make_item(
                workload_type="FMAPI_DATABRICKS",
                workload_name="Token workload",
                fmapi_model="llama-3-3-70b",
                fmapi_rate_type="input_token",
                fmapi_quantity=100,
            ),
            ("N", "O"),
        ),
    ],
)
def test_non_quantity_exports_keep_live_dbu_formulas(item, expected_columns):
    output = build_estimate_excel(
        make_estimate(),
        [item],
        cloud="aws",
        region="us-east-1",
        tier="PREMIUM",
    )
    formulas = openpyxl.load_workbook(output, data_only=False).active

    row = find_row_by_name(formulas, item.workload_name)
    assert row is not None
    assert formulas.cell(row, COL_DBUS_MO).value == (
        f"={expected_columns[0]}{row}*{expected_columns[1]}{row}"
    )
