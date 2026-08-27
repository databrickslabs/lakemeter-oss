"""General Storage pricing, persistence, parity, and export regressions."""
from datetime import datetime, timezone
from io import BytesIO
import json
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import openpyxl
import pytest
from fastapi import HTTPException

from app.models.line_item import LineItem
from app.routes.calculate import general_storage_calc
from app.routes.calculate.general_storage_calc import (
    calculate_general_storage_cost,
)
from app.routes.calculate.helpers import get_sku_type as get_calculation_sku
from app.routes.calculate.schemas import GeneralStorageCalculationRequest
from app.routes.estimates import _copy_line_item
from app.routes.export.excel_builder import build_estimate_excel
from app.routes.export.pricing import _get_sku_type as get_export_sku
from app.routes.line_items import _copy_line_item_for_clone
from app.routes.workload_types import DEFAULT_WORKLOAD_TYPES
from app.schemas.line_item import (
    GENERAL_STORAGE_CONFIG_FIELDS,
    LineItemCreate,
    LineItemResponse,
    LineItemUpdate,
    map_ai_parse_api_fields,
    validate_general_storage_workload_config,
)
from app.services.ai_agent import EstimateAgent, SYSTEM_PROMPT, TOOLS
from app.services.general_storage_pricing import (
    GENERAL_STORAGE_OPERATION_DSU_RATES,
    GENERAL_STORAGE_SKU,
    calculate_general_storage_usage,
)
from app.services.lakebase_queries import get_sku_type as get_service_sku
from tests.parity.frontend_calc import fe_general_storage_cost
from tests.regression.conftest import make_item
from tests.regression.excel_helpers import (
    COL_CONFIG,
    COL_DBU_COST_L,
    COL_DBU_HR,
    COL_DBU_RATE,
    COL_DBUS_MO,
    COL_DSU_COST_D,
    COL_DSU_COST_L,
    COL_DSU_RATE,
    COL_DSUS_MO,
    COL_DRIVER,
    COL_HOURS,
    COL_NOTES,
    COL_SKU,
    COL_TOTAL_L,
    COL_TOTAL_VM,
    COL_WORKER,
    find_row_by_name,
    make_estimate,
)


ROOT = Path(__file__).resolve().parents[2]


class _Result:
    def __init__(self, row):
        self.row = row

    def fetchone(self):
        return self.row


class _PricingDb:
    def __init__(self, price=0.023):
        self.price = price

    def execute(self, _statement, params):
        assert params["pt"] == GENERAL_STORAGE_SKU
        row = (
            SimpleNamespace(price_per_dbu=self.price)
            if self.price is not None
            else None
        )
        return _Result(row)


@pytest.fixture(autouse=True)
def _patch_reference_validation(monkeypatch):
    monkeypatch.setattr(
        general_storage_calc,
        "validate_cloud",
        lambda *args, **kwargs: None,
    )
    monkeypatch.setattr(
        general_storage_calc,
        "validate_region",
        lambda *args, **kwargs: None,
    )
    monkeypatch.setattr(
        general_storage_calc,
        "validate_tier",
        lambda *args, **kwargs: None,
    )


def _request(**overrides):
    values = {
        "cloud": "AWS",
        "region": "us-east-1",
        "tier": "PREMIUM",
        "quantity": 100,
        "unit": "gb",
        "tier_1_operations_thousands": 0,
        "tier_2_operations_thousands": 0,
    }
    values.update(overrides)
    return GeneralStorageCalculationRequest(**values)


class TestUsageAndCalculation:
    @pytest.mark.parametrize(
        ("quantity", "unit", "expected_gb"),
        [(100, "gb", 100), (1.5, "tb", 1536), (0, "gb", 0)],
    )
    def test_gb_tb_conversion(self, quantity, unit, expected_gb):
        usage = calculate_general_storage_usage(quantity, unit)
        assert usage["billable_gb_months"] == expected_gb
        assert usage["stored_data_dsu"] == expected_gb

    @pytest.mark.parametrize(
        ("cloud", "tier_1_rate", "tier_2_rate"),
        [
            ("aws", 0.2174, 0.0174),
            ("azure", 0.3535, 0.0226),
            ("gcp", 0.2174, 0.0174),
        ],
    )
    def test_cloud_operation_dsu_matrix(
        self,
        cloud,
        tier_1_rate,
        tier_2_rate,
    ):
        usage = calculate_general_storage_usage(
            100,
            "gb",
            cloud,
            10,
            20,
        )
        assert usage["tier_1_operations_dsu"] == pytest.approx(
            10 * tier_1_rate
        )
        assert usage["tier_2_operations_dsu"] == pytest.approx(
            20 * tier_2_rate
        )
        assert usage["total_dsu"] == pytest.approx(
            100 + 10 * tier_1_rate + 20 * tier_2_rate
        )
        assert (
            GENERAL_STORAGE_OPERATION_DSU_RATES[cloud][
                "tier_1_per_thousand"
            ]
            == tier_1_rate
        )

    @pytest.mark.parametrize(
        ("quantity", "unit"),
        [(-1, "gb"), (float("inf"), "gb"), (1, "pb")],
    )
    def test_invalid_usage_is_rejected(self, quantity, unit):
        with pytest.raises(ValueError):
            calculate_general_storage_usage(quantity, unit)

    def test_api_uses_exact_regional_rate_and_dsu_components(self):
        data = calculate_general_storage_cost(
            _request(
                quantity=2,
                unit="tb",
                tier_1_operations_thousands=1000,
                tier_2_operations_thousands=500,
            ),
            db=_PricingDb(price=0.041),
        )["data"]
        assert data["sku_type"] == GENERAL_STORAGE_SKU
        assert data["usage_calculation"]["billable_gb_months"] == 2048
        expected_dsu = 2048 + 1000 * 0.2174 + 500 * 0.0174
        assert data["total_cost"]["cost_per_month"] == pytest.approx(
            expected_dsu * 0.041
        )
        assert data["dsu_calculation"]["total_dsu"] == pytest.approx(
            expected_dsu
        )
        assert len(data["sku_breakdown"]) == 3
        assert {
            line["rate_type"] for line in data["sku_breakdown"]
        } == {
            "stored_data",
            "tier_1_operations",
            "tier_2_operations",
        }
        assert data["sku_breakdown"][0]["type"] == "dsu"
        assert data["sku_breakdown"][0]["usage_unit"] == "DSU"
        assert data["sku_breakdown"][0]["qty"] == 2048

    def test_zero_usage_is_valid(self):
        data = calculate_general_storage_cost(
            _request(quantity=0),
            db=_PricingDb(),
        )["data"]
        assert data["total_cost"]["cost_per_month"] == 0
        assert len(data["sku_breakdown"]) == 3
        assert all(line["qty"] == 0 for line in data["sku_breakdown"])

    def test_missing_exact_price_is_rejected(self):
        with pytest.raises(HTTPException, match="pricing is not available"):
            calculate_general_storage_cost(
                _request(),
                db=_PricingDb(price=None),
            )

    def test_frontend_backend_formula_parity(self):
        data = calculate_general_storage_cost(
            _request(
                quantity=1.25,
                unit="tb",
                tier_1_operations_thousands=200,
                tier_2_operations_thousands=300,
            ),
            db=_PricingDb(price=0.025),
        )["data"]
        assert data["total_cost"]["cost_per_month"] == pytest.approx(
            fe_general_storage_cost(
                quantity=1.25,
                unit="tb",
                cloud="aws",
                tier1_operations_thousands=200,
                tier2_operations_thousands=300,
                price_per_dsu=0.025,
            )
        )

    def test_storage_discount_applies_to_all_dsu_components(self):
        data = calculate_general_storage_cost(
            _request(
                quantity=100,
                tier_1_operations_thousands=10,
                tier_2_operations_thousands=20,
                discount_config={
                    "global": {"storage_discount": 10},
                },
            ),
            db=_PricingDb(price=0.023),
        )["data"]
        assert data["total_cost"]["total_after_discount"] == pytest.approx(
            sum(
                line["cost_after_discount"]
                for line in data["sku_breakdown"]
            )
        )
        assert all(
            line["discount"]["source"] == "global:storage"
            for line in data["sku_breakdown"]
        )


class TestPersistenceAndRegistries:
    fields = {
        "general_storage_quantity": 2.5,
        "general_storage_unit": "tb",
        "general_storage_tier1_operations_thousands": 125,
        "general_storage_tier2_operations_thousands": 250,
    }

    def test_create_update_response_and_type_cleanup(self):
        item = LineItemCreate.model_validate({
            "estimate_id": uuid4(),
            "workload_name": "Default Storage",
            "workload_type": "GENERAL_STORAGE",
            **self.fields,
        })
        mapped = map_ai_parse_api_fields(
            item.model_dump(),
            item.model_fields_set,
        )
        assert mapped["workload_config"] == self.fields
        assert not (set(GENERAL_STORAGE_CONFIG_FIELDS) & set(mapped))
        validate_general_storage_workload_config(
            mapped["workload_type"],
            mapped["workload_config"],
        )
        assert not hasattr(LineItem, "general_storage_quantity")
        LineItem(**mapped)

        update = LineItemUpdate(general_storage_quantity=3)
        updated = map_ai_parse_api_fields(
            update.model_dump(exclude_unset=True),
            update.model_fields_set,
            existing_workload_config=mapped["workload_config"],
        )
        assert updated["workload_config"] == {
            "general_storage_quantity": 3,
            "general_storage_unit": "tb",
            "general_storage_tier1_operations_thousands": 125,
            "general_storage_tier2_operations_thousands": 250,
        }
        now = datetime.now(timezone.utc)
        response = LineItemResponse.model_validate({
            "line_item_id": uuid4(),
            "estimate_id": uuid4(),
            "workload_name": "Default Storage",
            "workload_type": "GENERAL_STORAGE",
            "workload_config": updated["workload_config"],
            "created_at": now,
            "updated_at": now,
        })
        assert response.general_storage_quantity == 3
        assert response.general_storage_unit == "tb"
        assert (
            response.general_storage_tier1_operations_thousands == 125
        )
        assert (
            response.general_storage_tier2_operations_thousands == 250
        )

        cleanup = LineItemUpdate(workload_type="JOBS")
        cleaned = map_ai_parse_api_fields(
            cleanup.model_dump(exclude_unset=True),
            cleanup.model_fields_set,
            existing_workload_config=updated["workload_config"],
        )
        assert cleaned["workload_config"] is None

    @pytest.mark.parametrize("clone_kind", ["estimate", "line_item"])
    def test_clones_deep_copy_workload_config(self, clone_kind):
        original = LineItem(
            estimate_id=uuid4(),
            workload_name="Default Storage",
            workload_type="GENERAL_STORAGE",
            workload_config={**self.fields, "nested": {"value": [1]}},
        )
        cloned = (
            _copy_line_item(original, uuid4())
            if clone_kind == "estimate"
            else _copy_line_item_for_clone(original, 2, "Storage Copy")
        )
        assert cloned.workload_config == original.workload_config
        assert cloned.workload_config is not original.workload_config

    def test_workload_sku_and_assistant_registries(self):
        workload = next(
            item
            for item in DEFAULT_WORKLOAD_TYPES
            if item["workload_type"] == "GENERAL_STORAGE"
        )
        assert workload["sku_product_type_standard"] == GENERAL_STORAGE_SKU
        assert get_calculation_sku("GENERAL_STORAGE") == GENERAL_STORAGE_SKU
        assert get_service_sku("GENERAL_STORAGE") == GENERAL_STORAGE_SKU
        assert get_export_sku(
            SimpleNamespace(workload_type="GENERAL_STORAGE"),
            "aws",
        ) == GENERAL_STORAGE_SKU

        tool = next(t for t in TOOLS if t["name"] == "propose_workload")
        properties = tool["parameters"]["properties"]
        assert "GENERAL_STORAGE" in properties["workload_type"]["enum"]
        assert "exact regional DATABRICKS_STORAGE" in SYSTEM_PROMPT
        agent = EstimateAgent(None)
        agent.current_estimate = {
            "cloud": "aws",
            "region": "us-east-1",
            "tier": "PREMIUM",
        }
        proposal = agent._propose_workload(
            "GENERAL_STORAGE",
            "Default Storage",
        )["proposed_workload"]
        assert proposal["general_storage_quantity"] == 100
        assert proposal["general_storage_unit"] == "gb"
        assert proposal["general_storage_tier1_operations_thousands"] == 0
        assert proposal["general_storage_tier2_operations_thousands"] == 0

    def test_all_published_regions_have_storage_prices(self):
        rates = json.loads(
            (
                ROOT / "backend/static/pricing/dbu-rates.json"
            ).read_text(encoding="utf-8")
        )
        assert rates["aws:eu-north-1:PREMIUM"][GENERAL_STORAGE_SKU] == 0.023
        assert (
            rates["gcp:southamerica-east1:ENTERPRISE"][
                GENERAL_STORAGE_SKU
            ]
            == 0.041
        )


class TestExcelExport:
    def test_storage_row_has_exact_sku_rate_formula_and_no_compute(self):
        item = make_item(
            workload_type="GENERAL_STORAGE",
            workload_name="Default Storage",
            workload_config={
                "general_storage_quantity": 2,
                "general_storage_unit": "tb",
                "general_storage_tier1_operations_thousands": 1000,
                "general_storage_tier2_operations_thousands": 500,
            },
        )
        output = build_estimate_excel(
            make_estimate(),
            [item],
            "aws",
            "us-east-1",
            "PREMIUM",
        )
        payload = BytesIO(output.getvalue())
        formulas = openpyxl.load_workbook(
            payload,
            data_only=False,
        ).active
        payload.seek(0)
        cached = openpyxl.load_workbook(payload, data_only=True).active
        rows = [
            row
            for row in range(1, formulas.max_row + 1)
            if str(formulas.cell(row, 2).value or "").startswith(
                "Default Storage – "
            )
        ]
        assert len(rows) == 3
        expected_dsus = [2048, 217.4, 8.7]
        for row, expected_dsu in zip(rows, expected_dsus):
            assert formulas.cell(row, COL_SKU).value == GENERAL_STORAGE_SKU
            assert formulas.cell(row, COL_DBU_RATE).value == 0
            assert formulas.cell(row, COL_DSU_RATE).value == 0.023
            assert formulas.cell(row, COL_HOURS).value == "N/A"
            assert formulas.cell(row, COL_DBU_HR).value == "N/A"
            assert formulas.cell(row, COL_DBUS_MO).value == 0
            assert formulas.cell(row, COL_DRIVER).value == "-"
            assert formulas.cell(row, COL_WORKER).value == "-"
            assert cached.cell(row, COL_TOTAL_VM).value == 0
            assert cached.cell(row, COL_DBU_COST_L).value == 0
            assert cached.cell(row, COL_DSUS_MO).value == pytest.approx(
                expected_dsu
            )
            assert cached.cell(row, COL_DSU_COST_L).value == pytest.approx(
                expected_dsu * 0.023
            )
            assert cached.cell(row, COL_DSU_COST_D).value == pytest.approx(
                expected_dsu * 0.023
            )
            assert cached.cell(row, COL_TOTAL_L).value == pytest.approx(
                expected_dsu * 0.023
            )
        assert "2048 GB-month × 1 DSU/GB-month" in formulas.cell(
            rows[0],
            COL_CONFIG,
        ).value
        assert "0.2174 DSU/1K" in formulas.cell(
            rows[1],
            COL_CONFIG,
        ).value
        notes = formulas.cell(rows[-1], COL_NOTES).value
        assert "$0.023/DSU" in notes
        assert "customer-managed object storage" in notes

    def test_export_rejects_unknown_region_without_fallback(self):
        item = make_item(
            workload_type="GENERAL_STORAGE",
            workload_name="Default Storage",
            workload_config=self._storage_config(),
        )
        with pytest.raises(ValueError, match="pricing is not available"):
            build_estimate_excel(
                make_estimate(),
                [item],
                "aws",
                "not-a-region",
                "PREMIUM",
            )

    @staticmethod
    def _storage_config():
        return {
            "general_storage_quantity": 100,
            "general_storage_unit": "gb",
            "general_storage_tier1_operations_thousands": 0,
            "general_storage_tier2_operations_thousands": 0,
        }
