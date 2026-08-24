"""Regression tests for the AI Extract and AI Classify workload types.

Covers the acceptance criteria of the feature request:
  - calculation endpoints price presets and custom rates against the
    SERVERLESS_REAL_TIME_INFERENCE SKU
  - every SKU resolver (Python and SQL) maps both types
  - line-item fields persist, clone, and round-trip without alias mapping
  - Excel export DBU quantities match the calculation API
  - exported DBUs/Mo cells are written as values, so a full recalculation
    (Google Sheets, LibreOffice, Ctrl+Alt+F9) cannot collapse them to zero
"""
from datetime import datetime, timezone
import io
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import pytest

from app.routes.calculate import ai_extract_calc, ai_classify_calc
from app.routes.calculate.ai_extract_calc import (
    EXTRACT_DOCUMENT_RATES, calculate_ai_extract_cost,
)
from app.routes.calculate.ai_classify_calc import (
    CLASSIFY_DOCUMENT_RATES, calculate_ai_classify_cost,
)
from app.routes.calculate.helpers import get_sku_type as get_calculation_sku_type
from app.routes.calculate.schemas import (
    AIExtractCalculationRequest, AIClassifyCalculationRequest,
)
from app.routes.export.excel_builder import build_estimate_excel
from app.routes.export.excel_item_helpers import calc_item_values
from app.routes.export.pricing import _get_sku_type as get_export_sku_type
from app.routes.estimates import _copy_line_item
from app.routes.line_items import _copy_line_item_for_clone
from app.models.line_item import LineItem
from app.schemas.line_item import (
    AI_FUNCTION_CONFIG_FIELDS,
    LineItemCreate,
    LineItemResponse,
    LineItemUpdate,
    map_ai_parse_api_fields,
    validate_ai_function_workload_config,
)
from app.services.ai_agent import EstimateAgent, TOOLS
from app.services.lakebase_queries import get_sku_type as get_service_sku_type

RTI_PRICE = 0.07  # AWS us-east-1 PREMIUM reference rate for the SRTI SKU


class _Result:
    def __init__(self, row):
        self._row = row

    def fetchone(self):
        return self._row


class _PricingDb:
    def __init__(self, price=RTI_PRICE):
        self.price = price
        self.product_types = []

    def execute(self, statement, params):
        self.product_types.append(params["pt"])
        row = (
            SimpleNamespace(price_per_dbu=self.price)
            if self.price is not None
            else None
        )
        return _Result(row)


@pytest.fixture(autouse=True)
def _patch_validation(monkeypatch):
    for module in (ai_extract_calc, ai_classify_calc):
        monkeypatch.setattr(module, "validate_cloud", lambda *a, **k: None)
        monkeypatch.setattr(module, "validate_region", lambda *a, **k: None)
        monkeypatch.setattr(module, "validate_tier", lambda *a, **k: None)


def _extract_request(**kwargs):
    base = dict(cloud="AWS", region="us-east-1", tier="PREMIUM")
    base.update(kwargs)
    return AIExtractCalculationRequest(**base)


def _classify_request(**kwargs):
    base = dict(cloud="AWS", region="us-east-1", tier="PREMIUM")
    base.update(kwargs)
    return AIClassifyCalculationRequest(**base)


class TestCalculation:
    @pytest.mark.parametrize(
        ("document_type", "num_inputs", "expected_dbu"),
        [
            ("short_text", 100_000, 4_500.0),
            ("invoice", 100_000, 4_500.0),
            ("complex_reasoning", 10_000, 5_625.0),
            ("deep_nesting", 10_000, 5_375.0),
        ],
    )
    def test_extract_presets(self, document_type, num_inputs, expected_dbu):
        response = calculate_ai_extract_cost(
            _extract_request(document_type=document_type, num_inputs=num_inputs),
            db=_PricingDb(),
        )
        data = response["data"]
        assert data["sku_type"] == "SERVERLESS_REAL_TIME_INFERENCE"
        assert data["dbu_calculation"]["dbu_per_month"] == expected_dbu
        assert data["total_cost"]["cost_per_month"] == pytest.approx(
            expected_dbu * RTI_PRICE)
        assert any("ai_parse_document" in note for note in data["notes"])

    @pytest.mark.parametrize(
        ("document_type", "num_docs", "expected_dbu"),
        [
            ("short_text", 500_000, 2_250.0),
            ("rental_contract", 20_000, 1_000.0),
        ],
    )
    def test_classify_presets(self, document_type, num_docs, expected_dbu):
        response = calculate_ai_classify_cost(
            _classify_request(document_type=document_type, num_docs=num_docs),
            db=_PricingDb(),
        )
        data = response["data"]
        assert data["sku_type"] == "SERVERLESS_REAL_TIME_INFERENCE"
        assert data["dbu_calculation"]["dbu_per_month"] == expected_dbu
        assert data["total_cost"]["cost_per_month"] == pytest.approx(
            expected_dbu * RTI_PRICE)
        assert any("ai_parse_document" in note for note in data["notes"])

    def test_extract_custom_rate(self):
        response = calculate_ai_extract_cost(
            _extract_request(document_type="custom", num_inputs=5_000,
                             dbus_per_thousand=80),
            db=_PricingDb(),
        )
        assert response["data"]["dbu_calculation"]["dbu_per_month"] == 400.0

    def test_classify_custom_rate(self):
        response = calculate_ai_classify_cost(
            _classify_request(document_type="custom", num_docs=5_000,
                              dbus_per_thousand=8),
            db=_PricingDb(),
        )
        assert response["data"]["dbu_calculation"]["dbu_per_month"] == 40.0

    def test_extract_rejects_unknown_document_type(self):
        from fastapi import HTTPException
        with pytest.raises(HTTPException) as exc_info:
            calculate_ai_extract_cost(
                _extract_request(document_type="novel"), db=_PricingDb())
        assert exc_info.value.status_code == 400

    def test_custom_requires_rate(self):
        from fastapi import HTTPException
        with pytest.raises(HTTPException) as exc_info:
            calculate_ai_classify_cost(
                _classify_request(document_type="custom", num_docs=1_000),
                db=_PricingDb())
        assert exc_info.value.status_code == 400

    @pytest.mark.parametrize(
        ("calculator", "calculation_request"),
        [
            (
                calculate_ai_extract_cost,
                _extract_request(document_type="invoice", num_inputs=1_000),
            ),
            (
                calculate_ai_classify_cost,
                _classify_request(document_type="short_text", num_docs=1_000),
            ),
        ],
    )
    def test_missing_regional_price_is_rejected(
        self,
        calculator,
        calculation_request,
    ):
        from fastapi import HTTPException

        with pytest.raises(HTTPException) as exc_info:
            calculator(calculation_request, db=_PricingDb(price=None))

        assert exc_info.value.status_code == 400
        assert "pricing is not available" in exc_info.value.detail


class TestSkuResolution:
    @pytest.mark.parametrize("workload_type", ["AI_EXTRACT", "AI_CLASSIFY"])
    def test_python_resolvers(self, workload_type):
        assert get_calculation_sku_type(workload_type, False, False) == \
            "SERVERLESS_REAL_TIME_INFERENCE"
        assert get_service_sku_type(workload_type, False, False) == \
            "SERVERLESS_REAL_TIME_INFERENCE"
        item = SimpleNamespace(workload_type=workload_type)
        assert get_export_sku_type(item, "aws") == "SERVERLESS_REAL_TIME_INFERENCE"

    @pytest.mark.parametrize("relative_path", [
        "scripts/functions/01_Utility_Functions.py",
        "etl/lakebase_setup/functions/01_Utility_Functions.py",
    ])
    def test_sql_resolvers(self, relative_path):
        repository_root = Path(__file__).resolve().parents[2]
        source = (repository_root / relative_path).read_text(encoding="utf-8")
        for workload_type in ("AI_EXTRACT", "AI_CLASSIFY"):
            assert f"WHEN '{workload_type}' THEN" in source, (
                f"{relative_path} is missing the {workload_type} mapping")


class TestInstallerCompatibility:
    @pytest.mark.parametrize("relative_path", [
        "scripts/install_lakemeter.py",
        "scripts/notebooks/02_create_database.py",
    ])
    def test_installer_seeds_types_without_schema_columns(self, relative_path):
        repository_root = Path(__file__).resolve().parents[2]
        source = (repository_root / relative_path).read_text(encoding="utf-8")

        assert '"AI_EXTRACT"' in source
        assert '"AI_CLASSIFY"' in source
        for field in AI_FUNCTION_CONFIG_FIELDS:
            assert f"{field} VARCHAR" not in source
            assert f'{field}", "NUMERIC' not in source


class TestAssistantPayload:
    def test_tool_schema_includes_all_json_backed_fields(self):
        propose_tool = next(
            tool for tool in TOOLS if tool["name"] == "propose_workload"
        )
        properties = propose_tool["parameters"]["properties"]

        assert set(AI_FUNCTION_CONFIG_FIELDS).issubset(properties)
        assert properties["ai_extract_document_type"]["enum"] == [
            "short_text",
            "invoice",
            "complex_reasoning",
            "deep_nesting",
            "custom",
        ]
        assert properties["ai_classify_document_type"]["enum"] == [
            "short_text",
            "rental_contract",
            "custom",
        ]
        assert properties["ai_extract_dbus_per_thousand"][
            "exclusiveMinimum"
        ] == 0
        assert properties["ai_classify_dbus_per_thousand"][
            "exclusiveMinimum"
        ] == 0

    def test_proposal_preserves_ai_function_payload(self):
        agent = EstimateAgent(None)
        agent.current_estimate = {
            "cloud": "aws",
            "region": "us-east-1",
            "tier": "PREMIUM",
        }
        payload = {
            "ai_extract_document_type": "complex_reasoning",
            "ai_extract_num_inputs": 10_000,
            "ai_extract_dbus_per_thousand": 600,
        }

        result = agent._propose_workload(
            "AI_EXTRACT",
            "Precision extraction",
            **payload,
        )

        assert result["success"] is True
        proposal = result["proposed_workload"]
        for field, value in payload.items():
            assert proposal[field] == value


class TestPersistenceAndClone:
    EXTRACT_FIELDS = {
        "ai_extract_document_type": "complex_reasoning",
        "ai_extract_num_inputs": 10_000,
        "ai_extract_dbus_per_thousand": 80,
    }
    CLASSIFY_FIELDS = {
        "ai_classify_document_type": "rental_contract",
        "ai_classify_num_docs": 20_000,
        "ai_classify_dbus_per_thousand": 8,
    }

    def test_line_item_create_maps_fields_to_workload_config(self):
        item = LineItemCreate.model_validate({
            "estimate_id": uuid4(),
            "workload_name": "extraction",
            "workload_type": "AI_EXTRACT",
            **self.EXTRACT_FIELDS,
        })
        mapped = map_ai_parse_api_fields(
            item.model_dump(),
            item.model_fields_set,
        )

        assert mapped["workload_config"] == self.EXTRACT_FIELDS
        assert not (set(AI_FUNCTION_CONFIG_FIELDS) & mapped.keys())
        assert not any(hasattr(LineItem, field) for field in self.EXTRACT_FIELDS)
        LineItem(**mapped)

    def test_line_item_update_merges_and_clears_json_fields(self):
        existing = {
            **self.EXTRACT_FIELDS,
            "unrelated_setting": "preserved",
        }
        update = LineItemUpdate(
            ai_extract_num_inputs=25_000,
            ai_extract_dbus_per_thousand=None,
        )
        mapped = map_ai_parse_api_fields(
            update.model_dump(exclude_unset=True),
            update.model_fields_set,
            existing_workload_config=existing,
        )

        assert mapped["workload_config"] == {
            "ai_extract_document_type": "complex_reasoning",
            "ai_extract_num_inputs": 25_000,
            "unrelated_setting": "preserved",
        }
        assert not (set(AI_FUNCTION_CONFIG_FIELDS) & mapped.keys())

    def test_response_hydrates_json_backed_fields(self):
        now = datetime.now(timezone.utc)
        response = LineItemResponse.model_validate({
            "line_item_id": uuid4(),
            "estimate_id": uuid4(),
            "workload_name": "classification",
            "workload_type": "AI_CLASSIFY",
            "workload_config": self.CLASSIFY_FIELDS,
            "created_at": now,
            "updated_at": now,
        })

        for key, value in self.CLASSIFY_FIELDS.items():
            assert getattr(response, key) == value

    def test_switching_workload_type_clears_irrelevant_ai_config(self):
        update = LineItemUpdate(workload_type="JOBS")
        mapped = map_ai_parse_api_fields(
            update.model_dump(exclude_unset=True),
            update.model_fields_set,
            existing_workload_config={
                **self.EXTRACT_FIELDS,
                "unrelated_setting": "preserved",
            },
        )

        assert mapped["workload_config"] == {
            "unrelated_setting": "preserved",
        }

    def test_custom_rate_validation_uses_merged_json(self):
        with pytest.raises(ValueError, match="required"):
            validate_ai_function_workload_config(
                "AI_EXTRACT",
                {
                    "ai_extract_document_type": "custom",
                    "ai_extract_num_inputs": 1000,
                },
            )
        with pytest.raises(ValueError, match="greater than 0"):
            validate_ai_function_workload_config(
                "AI_CLASSIFY",
                {
                    "ai_classify_document_type": "custom",
                    "ai_classify_num_docs": 1000,
                    "ai_classify_dbus_per_thousand": 0,
                },
            )

    def test_estimate_clone_copies_workload_config(self):
        original = LineItem(
            estimate_id=uuid4(), workload_name="extraction",
            workload_type="AI_EXTRACT",
            workload_config=self.EXTRACT_FIELDS,
        )
        new_estimate_id = uuid4()
        copy = _copy_line_item(original, new_estimate_id)

        assert copy.estimate_id == new_estimate_id
        assert copy.workload_config == self.EXTRACT_FIELDS
        assert copy.workload_config is not original.workload_config
        assert copy.line_item_id != original.line_item_id or copy.line_item_id is None

    def test_individual_clone_copies_workload_config(self):
        original = LineItem(
            estimate_id=uuid4(),
            workload_name="classification",
            workload_type="AI_CLASSIFY",
            workload_config=self.CLASSIFY_FIELDS,
        )
        copy = _copy_line_item_for_clone(original, 5, "classification (Copy)")

        assert copy.estimate_id == original.estimate_id
        assert copy.display_order == 5
        assert copy.workload_name == "classification (Copy)"
        assert copy.workload_config == self.CLASSIFY_FIELDS
        assert copy.workload_config is not original.workload_config


def _export_item(**kwargs):
    base = dict(
        workload_name="w", workload_type="AI_EXTRACT", serverless_enabled=True,
        photon_enabled=False, driver_node_type=None, worker_node_type=None,
        num_workers=None, hours_per_month=None, runs_per_day=None,
        avg_runtime_minutes=None, days_per_month=None, hours_per_day=None,
        driver_pricing_tier=None, worker_pricing_tier=None,
        driver_payment_option=None, worker_payment_option=None,
        dlt_edition=None, dbsql_warehouse_type=None, dbsql_warehouse_size=None,
        dbsql_num_clusters=None, dbsql_serverless_size=None,
        dbsql_vm_pricing_tier=None, dbsql_vm_payment_option=None,
        notes=None, serverless_mode=None, workload_config=None,
        ai_parse_calculation_method=None, ai_parse_complexity=None,
        ai_parse_num_pages=None, ai_parse_dbus=None, ai_parse_hours_per_month=None,
        ai_extract_document_type=None, ai_extract_num_inputs=None,
        ai_extract_dbus_per_thousand=None,
        ai_classify_document_type=None, ai_classify_num_docs=None,
        ai_classify_dbus_per_thousand=None,
        shutterstock_imageai_num_images=None,
        lakebase_cu=None, lakebase_ha_nodes=None, lakebase_storage_gb=None,
        lakebase_pitr_gb=None, lakebase_snapshot_gb=None,
        vector_search_storage_gb=None, vector_search_units=None,
        model_serving_concurrency=None, model_serving_gpu_type=None,
        fmapi_model=None, fmapi_provider=None, fmapi_input_tokens_millions=None,
        fmapi_output_tokens_millions=None, fmapi_provisioned_throughput=None,
        databricks_apps_size=None, databricks_apps_num_apps=None,
        databricks_support_tier=None,
    )
    base.update(kwargs)
    return SimpleNamespace(**base)


class TestExportParity:
    def test_calc_item_values_match_api(self):
        extract_item = _export_item(
            workload_type="AI_EXTRACT",
            ai_extract_document_type="invoice", ai_extract_num_inputs=100_000)
        _, _, _, total_dbus, _ = calc_item_values(
            extract_item, False, False, 0, "aws", [])
        assert total_dbus == 4_500.0

        classify_item = _export_item(
            workload_type="AI_CLASSIFY",
            ai_classify_document_type="rental_contract",
            ai_classify_num_docs=20_000)
        _, _, _, total_dbus, _ = calc_item_values(
            classify_item, False, False, 0, "aws", [])
        assert total_dbus == 1_000.0

    def test_custom_rate_export(self):
        item = _export_item(
            workload_type="AI_EXTRACT", ai_extract_document_type="custom",
            ai_extract_num_inputs=5_000, ai_extract_dbus_per_thousand=80)
        _, _, _, total_dbus, _ = calc_item_values(item, False, False, 0, "aws", [])
        assert total_dbus == 400.0

    def test_json_backed_fields_match_api(self):
        item = _export_item(
            workload_type="AI_CLASSIFY",
            workload_config={
                "ai_classify_document_type": "rental_contract",
                "ai_classify_num_docs": 20_000,
            },
        )
        _, _, _, total_dbus, _ = calc_item_values(
            item, False, False, 0, "aws", []
        )
        assert total_dbus == 1_000.0

    def test_export_calculation_rejects_non_positive_custom_rate(self):
        item = _export_item(
            workload_type="AI_EXTRACT",
            workload_config={
                "ai_extract_document_type": "custom",
                "ai_extract_num_inputs": 1_000,
                "ai_extract_dbus_per_thousand": 0,
            },
        )

        with pytest.raises(ValueError, match="greater than 0"):
            calc_item_values(item, False, False, 0, "aws", [])


class TestRecalculationSafeExport:
    def test_export_rejects_missing_regional_price(self):
        estimate = SimpleNamespace(
            estimate_name="AI Functions",
            cloud="AWS",
            region="unsupported-region",
            tier="PREMIUM",
            estimate_id="e1",
            status="draft",
        )
        item = _export_item(
            workload_name="extract",
            workload_type="AI_EXTRACT",
            workload_config={
                "ai_extract_document_type": "invoice",
                "ai_extract_num_inputs": 1_000,
            },
        )

        with pytest.raises(ValueError, match="pricing is not available"):
            build_estimate_excel(
                estimate,
                [item],
                "aws",
                "unsupported-region",
                "PREMIUM",
                db=None,
            )

    def test_dbu_cells_are_values_not_formulas(self):
        import openpyxl

        estimate = SimpleNamespace(
            estimate_name="AI Functions", cloud="AWS", region="us-east-1",
            tier="PREMIUM", estimate_id="e1", status="draft")
        items = [
            _export_item(workload_name="extract",
                         workload_type="AI_EXTRACT",
                         workload_config={
                             "ai_extract_document_type": "invoice",
                             "ai_extract_num_inputs": 100_000,
                         }),
            _export_item(workload_name="classify",
                         workload_type="AI_CLASSIFY",
                         workload_config={
                             "ai_classify_document_type": "short_text",
                             "ai_classify_num_docs": 500_000,
                         }),
        ]
        output = build_estimate_excel(estimate, items, "aws", "us-east-1",
                                      "PREMIUM", db=None)
        payload = io.BytesIO(output.getvalue())
        formulas = openpyxl.load_workbook(payload).active
        payload.seek(0)
        cached = openpyxl.load_workbook(payload, data_only=True).active

        expected = {"extract": 4_500.0, "classify": 2_250.0}
        seen = {}
        for row in range(1, formulas.max_row + 1):
            name = cached.cell(row=row, column=2).value
            if name in expected:
                dbu_cell = formulas.cell(row=row, column=17)
                assert not (isinstance(dbu_cell.value, str)
                            and dbu_cell.value.startswith("=")), (
                    f"{name}: DBUs/Mo must be a value, found formula "
                    f"{dbu_cell.value!r}")
                seen[name] = float(dbu_cell.value)
        assert seen == expected
