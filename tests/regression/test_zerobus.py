"""Regression coverage for standard and OpenTelemetry Zerobus Ingest."""
from types import SimpleNamespace
from uuid import uuid4

import openpyxl
import pytest

from app.routes.calculate import zerobus_calc
from app.routes.calculate.schemas import ZerobusCalculationRequest
from app.routes.calculate.zerobus_calc import calculate_zerobus_cost
from app.routes.export.excel_builder import build_estimate_excel
from app.routes.export.pricing import _get_sku_type
from app.routes.workload_types import DEFAULT_WORKLOAD_TYPES
from app.schemas.line_item import (
    LineItemCreate,
    LineItemResponse,
    map_ai_parse_api_fields,
    validate_zerobus_workload_config,
)
from app.services.zerobus_pricing import (
    ZEROBUS_DBU_PER_GB,
    calculate_zerobus_usage,
    validate_zerobus_availability,
)
from tests.regression.conftest import make_item
from tests.regression.excel_helpers import (
    COL_CONFIG,
    COL_DBU_RATE,
    COL_DBUS_MO,
    COL_SKU,
    find_row_by_name,
    make_estimate,
)


class _Result:
    def __init__(self, row):
        self.row = row

    def fetchone(self):
        return self.row


class _PricingDb:
    def __init__(self, price=0.35):
        self.price = price

    def execute(self, statement, params):
        if "pt" in params:
            return _Result(SimpleNamespace(price_per_dbu=self.price))
        if "sku" in params:
            return _Result(("dbu", True))
        raise AssertionError(f"Unexpected query params: {params}")


@pytest.fixture(autouse=True)
def _patch_reference_validation(monkeypatch):
    monkeypatch.setattr(
        zerobus_calc, "validate_cloud", lambda *args, **kwargs: None
    )
    monkeypatch.setattr(
        zerobus_calc, "validate_region", lambda *args, **kwargs: None
    )
    monkeypatch.setattr(
        zerobus_calc, "validate_tier", lambda *args, **kwargs: None
    )


@pytest.mark.parametrize(
    ("mode", "dbu_per_gb"),
    [("standard", 0.143), ("otel", 0.222)],
)
def test_usage_rates(mode, dbu_per_gb):
    usage = calculate_zerobus_usage(1000, mode)
    assert usage["dbu_per_gb"] == dbu_per_gb
    assert usage["monthly_dbus"] == pytest.approx(1000 * dbu_per_gb)
    assert ZEROBUS_DBU_PER_GB[mode] == dbu_per_gb


def test_mode_aliases_and_invalid_values():
    assert calculate_zerobus_usage(1, "normal")["mode"] == "standard"
    assert calculate_zerobus_usage(1, "opentelemetry")["mode"] == "otel"
    with pytest.raises(ValueError, match="standard or otel"):
        calculate_zerobus_usage(1, "unknown")
    with pytest.raises(ValueError, match="greater than or equal to 0"):
        calculate_zerobus_usage(-1, "standard")


@pytest.mark.parametrize(
    ("cloud", "tier"),
    [
        ("AWS", "PREMIUM"),
        ("AWS", "ENTERPRISE"),
        ("AZURE", "PREMIUM"),
        ("GCP", "PREMIUM"),
        ("GCP", "ENTERPRISE"),
    ],
)
def test_supported_cloud_tiers(cloud, tier):
    validate_zerobus_availability(cloud, tier)


@pytest.mark.parametrize(
    ("cloud", "tier"),
    [("AWS", "STANDARD"), ("AZURE", "ENTERPRISE"), ("GCP", "STANDARD")],
)
def test_unsupported_cloud_tiers(cloud, tier):
    with pytest.raises(ValueError, match="pricing is not available"):
        validate_zerobus_availability(cloud, tier)


def test_calculation_uses_jobs_serverless_price():
    request = ZerobusCalculationRequest(
        cloud="AWS",
        region="us-east-1",
        tier="PREMIUM",
        mode="otel",
        monthly_ingested_gb=1000,
    )
    data = calculate_zerobus_cost(request, db=_PricingDb())["data"]
    assert data["sku_type"] == "JOBS_SERVERLESS_COMPUTE"
    assert data["usage_calculation"]["monthly_dbus"] == pytest.approx(222)
    assert data["dbu_calculation"]["list_price_per_gb"] == pytest.approx(
        0.0777
    )
    assert data["total_cost"]["cost_per_month"] == pytest.approx(77.7)


def test_json_backed_persistence_round_trip():
    create = LineItemCreate(
        estimate_id=uuid4(),
        workload_name="Telemetry",
        workload_type="ZEROBUS",
        zerobus_mode="otel",
        zerobus_monthly_ingested_gb=250,
    )
    mapped = map_ai_parse_api_fields(
        create.model_dump(),
        create.model_fields_set,
    )
    assert mapped["workload_config"] == {
        "zerobus_mode": "otel",
        "zerobus_monthly_ingested_gb": 250,
    }
    validate_zerobus_workload_config(
        mapped["workload_type"],
        mapped["workload_config"],
    )
    response = LineItemResponse.model_validate({
        **mapped,
        "line_item_id": uuid4(),
        "created_at": "2026-08-29T00:00:00Z",
        "updated_at": "2026-08-29T00:00:00Z",
    })
    assert response.zerobus_mode == "otel"
    assert response.zerobus_monthly_ingested_gb == 250


def test_excel_exports_zerobus_usage_and_cost():
    item = make_item(
        workload_type="ZEROBUS",
        workload_name="Standard Zerobus",
        workload_config={
            "zerobus_mode": "standard",
            "zerobus_monthly_ingested_gb": 1000,
        },
    )
    output = build_estimate_excel(
        make_estimate(),
        [item],
        "aws",
        "us-east-1",
        "PREMIUM",
    )
    workbook = openpyxl.load_workbook(output, data_only=False)
    sheet = workbook["Databricks Estimate"]
    row = find_row_by_name(sheet, "Standard Zerobus")
    assert row is not None
    assert sheet.cell(row, COL_SKU).value == "JOBS_SERVERLESS_COMPUTE"
    assert sheet.cell(row, COL_DBUS_MO).value == pytest.approx(143)
    assert sheet.cell(row, COL_DBU_RATE).value == pytest.approx(0.35)
    assert "0.143 DBU/GB" in sheet.cell(row, COL_CONFIG).value


def test_workload_catalog_includes_zerobus():
    zerobus = next(
        item
        for item in DEFAULT_WORKLOAD_TYPES
        if item["workload_type"] == "ZEROBUS"
    )
    assert zerobus["sku_product_type_standard"] == "JOBS_SERVERLESS_COMPUTE"
    item = make_item(workload_type="ZEROBUS")
    assert _get_sku_type(item) == "JOBS_SERVERLESS_COMPUTE"
