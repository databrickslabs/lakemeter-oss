"""Focused AI Runtime pricing, persistence, and export regressions."""
from datetime import datetime, timezone
from io import BytesIO
import json
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import openpyxl
import pytest
from fastapi import HTTPException
from pydantic import ValidationError

from app.models.line_item import LineItem
from app.routes.calculate import ai_runtime_calc
from app.routes.calculate.ai_runtime_calc import calculate_ai_runtime_cost
from app.routes.calculate.helpers import get_sku_type as get_calculation_sku
from app.routes.calculate.schemas import AIRuntimeCalculationRequest
from app.routes.estimates import _copy_line_item
from app.routes.export.calculations import (
    _calculate_dbu_per_hour,
    _calculate_hours_per_month,
)
from app.routes.export.excel_builder import build_estimate_excel
from app.routes.export.pricing import _get_sku_type as get_export_sku
from app.routes.line_items import _copy_line_item_for_clone
from app.routes.workload_types import DEFAULT_WORKLOAD_TYPES
from app.schemas.line_item import (
    AI_RUNTIME_CONFIG_FIELDS,
    LineItemCreate,
    LineItemResponse,
    LineItemUpdate,
    map_ai_parse_api_fields,
    validate_ai_runtime_workload_config,
)
from app.services.ai_agent import EstimateAgent, SYSTEM_PROMPT, TOOLS
from app.services.ai_runtime_pricing import (
    AI_RUNTIME_ACCELERATORS,
    AI_RUNTIME_SKU,
    calculate_ai_runtime_usage,
    get_ai_runtime_accelerator,
)
from app.services.lakebase_queries import get_sku_type as get_service_sku
from tests.regression.conftest import make_item
from tests.regression.excel_helpers import (
    COL_CONFIG,
    COL_DBU_COST_L,
    COL_DBU_HR,
    COL_DBU_RATE,
    COL_DBUS_MO,
    COL_HOURS,
    COL_SKU,
    COL_TOTAL_L,
    find_row_by_name,
    make_estimate,
)


ROOT = Path(__file__).resolve().parents[2]


class _Result:
    def __init__(self, row):
        self.row = row

    def fetchone(self):
        return self.row


class _PricingDb:
    def __init__(self, price=0.65):
        self.price = price

    def execute(self, _statement, params):
        if "pt" in params:
            row = (
                SimpleNamespace(price_per_dbu=self.price)
                if self.price is not None
                else None
            )
            return _Result(row)
        if "sku" in params:
            return _Result(("dbu", False))
        raise AssertionError(f"Unexpected query params: {params}")


@pytest.fixture(autouse=True)
def _patch_reference_validation(monkeypatch):
    monkeypatch.setattr(
        ai_runtime_calc,
        "validate_cloud",
        lambda *args, **kwargs: None,
    )
    monkeypatch.setattr(
        ai_runtime_calc,
        "validate_region",
        lambda *args, **kwargs: None,
    )
    monkeypatch.setattr(
        ai_runtime_calc,
        "validate_tier",
        lambda *args, **kwargs: None,
    )


def _request(**overrides):
    values = {
        "cloud": "AWS",
        "region": "us-east-1",
        "tier": "PREMIUM",
        "accelerator_type": "GPU_1xA10",
        "hours_per_month": 10,
    }
    values.update(overrides)
    return AIRuntimeCalculationRequest(**values)


class TestPricing:
    @pytest.mark.parametrize(
        ("cloud", "accelerator", "gpu_count", "dbu_per_gpu_hour"),
        [
            ("aws", "GPU_1xA10", 1, 50 / 13),
            ("aws", "GPU_1xH100", 1, 140 / 13),
            ("aws", "GPU_8xH100", 8, 140 / 13),
            ("azure", "GPU_1xA10", 1, 98 / 13),
            ("azure", "GPU_1xH100", 1, 140 / 13),
            ("azure", "GPU_8xH100", 8, 140 / 13),
        ],
    )
    def test_accelerator_profiles(
        self,
        cloud,
        accelerator,
        gpu_count,
        dbu_per_gpu_hour,
    ):
        profile = get_ai_runtime_accelerator(cloud, accelerator)
        assert profile["gpu_count"] == gpu_count
        assert profile["dbu_per_gpu_hour"] == pytest.approx(
            dbu_per_gpu_hour
        )

    def test_profiles_preserve_published_us_east_list_prices(self):
        assert (
            AI_RUNTIME_ACCELERATORS["aws"]["GPU_1xA10"][
                "dbu_per_gpu_hour"
            ]
            * 0.65
        ) == pytest.approx(2.50)
        assert (
            AI_RUNTIME_ACCELERATORS["azure"]["GPU_1xA10"][
                "dbu_per_gpu_hour"
            ]
            * 0.65
        ) == pytest.approx(4.90)
        assert (
            AI_RUNTIME_ACCELERATORS["aws"]["GPU_1xH100"][
                "dbu_per_gpu_hour"
            ]
            * 0.65
        ) == pytest.approx(7.00)

    def test_eight_h100_counts_eight_gpu_hours_per_node_hour(self):
        usage = calculate_ai_runtime_usage(
            "aws",
            "GPU_8xH100",
            10,
        )
        assert usage["monthly_gpu_hours"] == 80
        assert usage["monthly_dbus"] == pytest.approx(
            10 * 8 * (140 / 13)
        )

    @pytest.mark.parametrize("cloud", ["aws", "azure"])
    def test_availability_is_not_restricted_by_region(self, cloud):
        usage = calculate_ai_runtime_usage(cloud, "GPU_1xH100", 1)
        assert usage["monthly_dbus"] == pytest.approx(140 / 13)

    def test_cloud_without_accelerator_rates_is_rejected(self):
        with pytest.raises(ValueError, match="supported on AWS and Azure"):
            get_ai_runtime_accelerator("gcp", "GPU_1xH100")


class TestCalculation:
    def test_direct_hours_use_model_training_sku_and_rate(self):
        data = calculate_ai_runtime_cost(
            _request(),
            db=_PricingDb(),
        )["data"]
        assert data["workload_type"] == "AI_RUNTIME"
        assert data["billing_origin_product"] == "AI_RUNTIME"
        assert data["sku_type"] == "MODEL_TRAINING"
        assert data["usage_calculation"]["monthly_gpu_hours"] == 10
        assert data["dbu_calculation"]["monthly_dbus"] == pytest.approx(
            10 * (50 / 13)
        )
        assert data["dbu_calculation"]["dbu_price"] == 0.65
        assert data["total_cost"]["cost_per_month"] == pytest.approx(25)

    def test_region_rollout_is_driven_by_exact_sku_price(self):
        data = calculate_ai_runtime_cost(
            _request(region="ap-southeast-1"),
            db=_PricingDb(price=0.85),
        )["data"]
        assert data["dbu_calculation"]["dbu_price"] == 0.85
        assert data["total_cost"]["cost_per_month"] == pytest.approx(
            10 * (50 / 13) * 0.85
        )

    def test_run_based_usage(self):
        data = calculate_ai_runtime_cost(
            _request(
                hours_per_month=None,
                runs_per_day=2,
                avg_runtime_minutes=90,
                days_per_month=20,
            ),
            db=_PricingDb(),
        )["data"]
        assert data["configuration"]["runtime_hours"] == 60
        assert data["total_cost"]["cost_per_month"] == pytest.approx(150)

    @pytest.mark.parametrize(
        ("cloud", "accelerator", "expected_cost"),
        [
            ("AWS", "GPU_1xH100", 70),
            ("AWS", "GPU_8xH100", 560),
            ("AZURE", "GPU_1xA10", 49),
        ],
    )
    def test_accelerator_costs(
        self,
        cloud,
        accelerator,
        expected_cost,
    ):
        region = "us-east-1" if cloud == "AWS" else "eastus"
        data = calculate_ai_runtime_cost(
            _request(
                cloud=cloud,
                region=region,
                accelerator_type=accelerator,
            ),
            db=_PricingDb(),
        )["data"]
        assert data["total_cost"]["cost_per_month"] == pytest.approx(
            expected_cost
        )

    def test_missing_exact_price_is_rejected(self):
        with pytest.raises(HTTPException, match="pricing is not available"):
            calculate_ai_runtime_cost(
                _request(),
                db=_PricingDb(price=None),
            )

    def test_standard_tier_is_rejected(self):
        with pytest.raises(HTTPException, match="Premium or Enterprise"):
            calculate_ai_runtime_cost(
                _request(tier="STANDARD"),
                db=_PricingDb(),
            )

    def test_cloud_without_accelerator_rates_is_rejected(self):
        with pytest.raises(
            HTTPException,
            match="supported on AWS and Azure",
        ):
            calculate_ai_runtime_cost(
                _request(
                    cloud="GCP",
                    region="us-central1",
                ),
                db=_PricingDb(),
            )

    @pytest.mark.parametrize(
        "overrides",
        [
            {"hours_per_month": -1},
            {"hours_per_month": float("inf")},
            {"hours_per_month": None},
            {"accelerator_type": "GPU_T4"},
        ],
    )
    def test_request_validation(self, overrides):
        with pytest.raises(ValidationError):
            _request(**overrides)


class TestPersistence:
    fields = {
        "ai_runtime_accelerator_type": "GPU_8xH100",
    }

    def test_create_folds_accelerator_into_json(self):
        item = LineItemCreate.model_validate({
            "estimate_id": uuid4(),
            "workload_name": "Training",
            "workload_type": "AI_RUNTIME",
            "hours_per_month": 10,
            **self.fields,
        })
        mapped = map_ai_parse_api_fields(
            item.model_dump(),
            item.model_fields_set,
        )
        validate_ai_runtime_workload_config(
            mapped["workload_type"],
            mapped["workload_config"],
        )
        assert mapped["workload_config"] == self.fields
        assert not (set(AI_RUNTIME_CONFIG_FIELDS) & set(mapped))
        assert not hasattr(LineItem, "ai_runtime_accelerator_type")
        LineItem(**mapped)

    def test_update_merges_and_response_hydrates(self):
        update = LineItemUpdate(
            ai_runtime_accelerator_type="GPU_1xH100",
        )
        mapped = map_ai_parse_api_fields(
            update.model_dump(exclude_unset=True),
            update.model_fields_set,
            existing_workload_config={
                **self.fields,
                "unrelated": {"nested": True},
            },
        )
        assert mapped["workload_config"] == {
            "ai_runtime_accelerator_type": "GPU_1xH100",
            "unrelated": {"nested": True},
        }
        now = datetime.now(timezone.utc)
        response = LineItemResponse.model_validate({
            "line_item_id": uuid4(),
            "estimate_id": uuid4(),
            "workload_name": "Training",
            "workload_type": "AI_RUNTIME",
            "workload_config": mapped["workload_config"],
            "created_at": now,
            "updated_at": now,
        })
        assert response.ai_runtime_accelerator_type == "GPU_1xH100"

    def test_switching_type_strips_runtime_fields(self):
        update = LineItemUpdate(workload_type="JOBS")
        mapped = map_ai_parse_api_fields(
            update.model_dump(exclude_unset=True),
            update.model_fields_set,
            existing_workload_config={
                **self.fields,
                "unrelated": {"nested": True},
            },
        )
        assert mapped["workload_config"] == {
            "unrelated": {"nested": True},
        }

    @pytest.mark.parametrize(
        "accelerator",
        [None, "GPU_T4", "gpu_1xa10"],
    )
    def test_json_validation_rejects_invalid_accelerators(
        self,
        accelerator,
    ):
        with pytest.raises(ValueError, match="must be one of"):
            validate_ai_runtime_workload_config(
                "AI_RUNTIME",
                {"ai_runtime_accelerator_type": accelerator},
            )

    @pytest.mark.parametrize("clone_kind", ["estimate", "line_item"])
    def test_clones_deep_copy_workload_config(self, clone_kind):
        original = LineItem(
            estimate_id=uuid4(),
            workload_name="Training",
            workload_type="AI_RUNTIME",
            workload_config={
                **self.fields,
                "unrelated": {"nested": ["value"]},
            },
        )
        if clone_kind == "estimate":
            cloned = _copy_line_item(original, uuid4())
        else:
            cloned = _copy_line_item_for_clone(
                original,
                2,
                "Training Copy",
            )
        assert cloned.workload_config == original.workload_config
        assert cloned.workload_config is not original.workload_config


class TestRegistriesAndAssistant:
    def test_workload_and_sku_registries(self):
        workload = next(
            item
            for item in DEFAULT_WORKLOAD_TYPES
            if item["workload_type"] == "AI_RUNTIME"
        )
        assert workload["display_name"] == "AI Runtime"
        assert workload["sku_product_type_standard"] == "MODEL_TRAINING"
        assert workload["display_order"] == 17
        assert get_calculation_sku("AI_RUNTIME") == AI_RUNTIME_SKU
        assert get_service_sku("AI_RUNTIME") == AI_RUNTIME_SKU
        assert get_export_sku(
            SimpleNamespace(workload_type="AI_RUNTIME"),
            "aws",
        ) == AI_RUNTIME_SKU
        for relative_path in (
            "scripts/functions/01_Utility_Functions.py",
            "etl/lakebase_setup/functions/01_Utility_Functions.py",
        ):
            source = (ROOT / relative_path).read_text(encoding="utf-8")
            assert "WHEN 'AI_RUNTIME' THEN" in source
            assert "v_product_type := 'MODEL_TRAINING'" in source

    def test_assistant_schema_defaults_prompt_and_mapping(self):
        tool = next(t for t in TOOLS if t["name"] == "propose_workload")
        properties = tool["parameters"]["properties"]
        assert "AI_RUNTIME" in properties["workload_type"]["enum"]
        assert properties["accelerator_type"]["enum"] == [
            "GPU_1xA10",
            "GPU_1xH100",
            "GPU_8xH100",
        ]
        assert "MODEL_TRAINING SKU" in SYSTEM_PROMPT

        agent = EstimateAgent(None)
        agent.current_estimate = {
            "cloud": "aws",
            "region": "us-east-1",
            "tier": "PREMIUM",
        }
        proposal = agent._propose_workload(
            "AI_RUNTIME",
            "Training",
            accelerator_type="GPU_1xH100",
            runs_per_day=2,
            avg_runtime_minutes=90,
            days_per_month=20,
        )["proposed_workload"]
        assert proposal["ai_runtime_accelerator_type"] == "GPU_1xH100"
        assert proposal["hours_per_month"] is None
        assert proposal["runs_per_day"] == 2

    @pytest.mark.parametrize(
        "relative_path",
        [
            "scripts/install_lakemeter.py",
            "scripts/notebooks/02_create_database.py",
        ],
    )
    def test_installer_seed_uses_json_without_schema_column(
        self,
        relative_path,
    ):
        source = (ROOT / relative_path).read_text(encoding="utf-8")
        assert '"AI_RUNTIME"' in source
        assert "ai_runtime_accelerator_type VARCHAR" not in source

    def test_frontend_tier_and_sku_availability_guards(self):
        form_source = (
            ROOT / "frontend/src/components/WorkloadForm.tsx"
        ).read_text(encoding="utf-8")
        premium_only = form_source.split(
            "const PREMIUM_ONLY_WORKLOAD_TYPES",
            maxsplit=1,
        )[1].split("])", maxsplit=1)[0]
        assert "'AI_RUNTIME'" in premium_only

        bundle_source = (
            ROOT / "frontend/src/utils/pricingBundle.ts"
        ).read_text(encoding="utf-8")
        assert "'MODEL_TRAINING': 'FMAPI_DATABRICKS'" not in bundle_source
        assert "AI_RUNTIME_SUPPORTED_REGIONS" not in bundle_source
        assert "productTypes.MODEL_TRAINING !== undefined" in bundle_source
        assert "getAIRuntimeAccelerators(cloudLower).length > 0" in bundle_source


class TestExcelExport:
    def test_direct_hours_export_matches_calculation(self):
        item = make_item(
            workload_type="AI_RUNTIME",
            workload_name="Training",
            hours_per_month=10,
            workload_config={
                "ai_runtime_accelerator_type": "GPU_1xH100",
            },
        )
        dbu_per_hour, warnings = _calculate_dbu_per_hour(
            item,
            "aws",
            "PREMIUM",
        )
        assert warnings == []
        assert dbu_per_hour == pytest.approx(140 / 13)
        assert _calculate_hours_per_month(item) == 10

        output = build_estimate_excel(
            make_estimate(),
            [item],
            "aws",
            "us-east-1",
            "PREMIUM",
        )
        payload = BytesIO(output.getvalue())
        formulas = openpyxl.load_workbook(
            payload,
            data_only=False,
        ).active
        payload.seek(0)
        cached = openpyxl.load_workbook(payload, data_only=True).active
        row = find_row_by_name(formulas, "Training")
        assert row is not None
        assert formulas.cell(row, COL_SKU).value == "MODEL_TRAINING"
        assert formulas.cell(row, COL_HOURS).value == 10
        assert formulas.cell(row, COL_DBU_HR).value == pytest.approx(
            140 / 13
        )
        assert cached.cell(row, COL_DBUS_MO).value == pytest.approx(
            10 * (140 / 13)
        )
        assert formulas.cell(row, COL_DBU_RATE).value == 0.65
        assert cached.cell(row, COL_DBU_COST_L).value == pytest.approx(70)
        assert cached.cell(row, COL_TOTAL_L).value == pytest.approx(70)
        assert "1x H100" in formulas.cell(row, COL_CONFIG).value

    def test_run_based_export_and_eight_gpu_multiplier(self):
        item = make_item(
            workload_type="AI_RUNTIME",
            workload_name="Distributed Training",
            runs_per_day=1,
            avg_runtime_minutes=120,
            days_per_month=5,
            workload_config={
                "ai_runtime_accelerator_type": "GPU_8xH100",
            },
        )
        output = build_estimate_excel(
            make_estimate(),
            [item],
            "aws",
            "us-east-1",
            "PREMIUM",
        )
        sheet = openpyxl.load_workbook(output, data_only=True).active
        row = find_row_by_name(sheet, "Distributed Training")
        assert row is not None
        assert sheet.cell(row, COL_HOURS).value == 10
        assert sheet.cell(row, COL_DBU_HR).value == pytest.approx(
            8 * (140 / 13)
        )
        assert sheet.cell(row, COL_TOTAL_L).value == pytest.approx(560)

    def test_export_uses_exact_price_without_region_allowlist(self):
        item = make_item(
            workload_type="AI_RUNTIME",
            workload_name="Training",
            hours_per_month=10,
            workload_config={
                "ai_runtime_accelerator_type": "GPU_1xA10",
            },
        )
        output = build_estimate_excel(
            make_estimate(),
            [item],
            "aws",
            "ap-southeast-1",
            "PREMIUM",
        )
        sheet = openpyxl.load_workbook(output, data_only=True).active
        row = find_row_by_name(sheet, "Training")
        assert row is not None
        assert sheet.cell(row, COL_DBU_RATE).value == 0.85
        assert sheet.cell(row, COL_TOTAL_L).value == pytest.approx(
            10 * (50 / 13) * 0.85
        )

    def test_export_rejects_standard_tier(self):
        item = make_item(
            workload_type="AI_RUNTIME",
            workload_name="Training",
            hours_per_month=10,
            workload_config={
                "ai_runtime_accelerator_type": "GPU_1xA10",
            },
        )
        with pytest.raises(ValueError, match="Premium or Enterprise"):
            build_estimate_excel(
                make_estimate(),
                [item],
                "aws",
                "us-east-1",
                "STANDARD",
            )


class TestReleaseWiring:
    def test_data_update_is_idempotent_and_has_no_migration(self):
        source = (
            ROOT
            / "scripts/upgrades/data_updates/024_ai_runtime.sql"
        ).read_text(encoding="utf-8")
        assert "'AI_RUNTIME'" in source
        assert "'MODEL_TRAINING'" in source
        assert "ON CONFLICT (workload_type) DO UPDATE" in source
        assert "ALTER TABLE" not in source.upper()

    def test_release_manifest_includes_ai_runtime_update(self):
        manifest = json.loads(
            (ROOT / "scripts/upgrades/release.json").read_text(
                encoding="utf-8"
            )
        )
        action = next(
            item
            for item in manifest["data_updates"]
            if item["id"] == "024-ai-runtime"
        )
        assert action["path"] == (
            "scripts/upgrades/data_updates/024_ai_runtime.sql"
        )
        assert len(action["sha256"]) == 64
