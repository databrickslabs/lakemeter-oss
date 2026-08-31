"""Focused Agent Evaluation pricing, persistence, and export regressions."""
from datetime import datetime, timezone
from io import BytesIO
import json
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import openpyxl
import pytest
from fastapi import HTTPException
from pydantic import ValidationError

from app.models.line_item import LineItem
from app.routes.calculate import agent_evaluation_calc
from app.routes.calculate.agent_evaluation_calc import (
    AGENT_EVALUATION_COMPONENT_RATES,
    calculate_agent_evaluation_cost,
    calculate_agent_evaluation_usage,
)
from app.routes.calculate.helpers import get_sku_type as get_calculation_sku
from app.routes.calculate.schemas import (
    AgentEvaluationCalculationRequest,
    DiscountConfig,
)
from app.routes.estimates import _copy_line_item
from app.routes.export.excel_builder import build_estimate_excel
from app.routes.export.excel_item_helpers import calc_item_values
from app.routes.export.pricing import _get_sku_type as get_export_sku
from app.routes.line_items import _copy_line_item_for_clone
from app.routes.workload_types import DEFAULT_WORKLOAD_TYPES
from app.schemas.line_item import (
    AGENT_EVALUATION_CONFIG_FIELDS,
    LineItemCreate,
    LineItemResponse,
    LineItemUpdate,
    map_ai_parse_api_fields,
    validate_agent_evaluation_workload_config,
)
from app.services.ai_agent import EstimateAgent, SYSTEM_PROMPT, TOOLS
from app.services.lakebase_queries import get_sku_type as get_service_sku
from tests.regression.conftest import make_item
from tests.regression.excel_helpers import (
    COL_CONFIG,
    COL_DBU_COST_L,
    COL_DBU_PER_M,
    COL_DBU_RATE,
    COL_DBUS_MO,
    COL_NOTES,
    COL_TOKEN_QTY,
    COL_TOTAL_L,
    find_row_by_name,
    find_totals_row,
    make_estimate,
)


ROOT = Path(__file__).resolve().parents[2]
SRTI_US_PRICE = 0.07
SRTI_SINGAPORE_PRICE = 0.088


class _Result:
    def __init__(self, row):
        self.row = row

    def fetchone(self):
        return self.row


class _PricingDb:
    def __init__(self, price=SRTI_US_PRICE):
        self.price = price

    def execute(self, _statement, params):
        if "pt" in params:
            row = (
                SimpleNamespace(price_per_dbu=self.price)
                if self.price is not None
                else None
            )
            return _Result(row)
        if "sku" in params:
            return _Result(("dbu", True))
        raise AssertionError(f"Unexpected query params: {params}")


@pytest.fixture(autouse=True)
def _patch_reference_validation(monkeypatch):
    monkeypatch.setattr(
        agent_evaluation_calc,
        "validate_cloud",
        lambda *args, **kwargs: None,
    )
    monkeypatch.setattr(
        agent_evaluation_calc,
        "validate_region",
        lambda *args, **kwargs: None,
    )
    monkeypatch.setattr(
        agent_evaluation_calc,
        "validate_tier",
        lambda *args, **kwargs: None,
    )


def _request(**kwargs):
    values = {
        "cloud": "AWS",
        "region": "us-east-1",
        "tier": "PREMIUM",
        "labels_enabled": True,
        "input_tokens_millions": 2,
        "output_tokens_millions": 3,
        "synthetic_data_enabled": True,
        "synthetic_questions": 4,
    }
    values.update(kwargs)
    return AgentEvaluationCalculationRequest(**values)


class TestCalculation:
    def test_canonical_component_rates(self):
        assert AGENT_EVALUATION_COMPONENT_RATES == {
            "input_tokens": 2.143,
            "output_tokens": 8.571,
            "synthetic_questions": 5.0,
        }
        usage = calculate_agent_evaluation_usage(
            labels_enabled=True,
            input_tokens_millions=2,
            output_tokens_millions=3,
            synthetic_data_enabled=True,
            synthetic_questions=4,
        )
        components = {
            component["component"]: component
            for component in usage["components"]
        }
        assert components["input_tokens"]["monthly_dbus"] == pytest.approx(
            4.286
        )
        assert components["output_tokens"]["monthly_dbus"] == pytest.approx(
            25.713
        )
        assert components["synthetic_questions"][
            "monthly_dbus"
        ] == pytest.approx(20)
        assert usage["monthly_dbus"] == pytest.approx(49.999)

    def test_singapore_uses_exact_regional_srti_price(self):
        data = calculate_agent_evaluation_cost(
            _request(region="ap-southeast-1"),
            db=_PricingDb(SRTI_SINGAPORE_PRICE),
        )["data"]
        assert data["dbu_calculation"]["dbu_price"] == 0.088
        assert data["dbu_calculation"]["monthly_dbu_cost"] == pytest.approx(
            49.999 * 0.088
        )
        assert data["total_cost"]["cost_per_month"] == pytest.approx(
            4.399912
        )
        assert len(data["sku_breakdown"]) == 1
        assert data["sku_breakdown"][0]["qty"] == pytest.approx(49.999)
        assert data["sku_type"] == "SERVERLESS_REAL_TIME_INFERENCE"

    def test_us_price_parity_reproduces_published_prices(self):
        data = calculate_agent_evaluation_cost(
            _request(
                input_tokens_millions=1,
                output_tokens_millions=1,
                synthetic_questions=1,
            ),
            db=_PricingDb(SRTI_US_PRICE),
        )["data"]
        costs = {
            component["component"]: component["monthly_dbu_cost"]
            for component in data["component_breakdown"]
        }
        assert costs["input_tokens"] == pytest.approx(0.15, abs=0.00002)
        assert costs["output_tokens"] == pytest.approx(0.60, abs=0.00004)
        assert costs["synthetic_questions"] == pytest.approx(0.35)

    @pytest.mark.parametrize(
        (
            "labels_enabled",
            "synthetic_enabled",
            "expected_components",
            "expected_dbus",
        ),
        [
            (
                True,
                False,
                ["input_tokens", "output_tokens"],
                29.999,
            ),
            (False, True, ["synthetic_questions"], 20),
            (
                True,
                True,
                [
                    "input_tokens",
                    "output_tokens",
                    "synthetic_questions",
                ],
                49.999,
            ),
        ],
    )
    def test_feature_toggles(
        self,
        labels_enabled,
        synthetic_enabled,
        expected_components,
        expected_dbus,
    ):
        usage = calculate_agent_evaluation_usage(
            labels_enabled=labels_enabled,
            input_tokens_millions=2,
            output_tokens_millions=3,
            synthetic_data_enabled=synthetic_enabled,
            synthetic_questions=4,
        )
        assert [
            component["component"] for component in usage["components"]
        ] == expected_components
        assert usage["monthly_dbus"] == pytest.approx(expected_dbus)

    def test_zero_values_remain_transparent_when_labels_enabled(self):
        data = calculate_agent_evaluation_cost(
            _request(
                input_tokens_millions=0,
                output_tokens_millions=0,
                synthetic_data_enabled=False,
                synthetic_questions=999,
            ),
            db=_PricingDb(),
        )["data"]
        assert list(data["usage_calculation"]) == [
            "input_tokens",
            "output_tokens",
        ]
        assert all(
            component["monthly_dbus"] == 0
            for component in data["component_breakdown"]
        )
        assert data["dbu_calculation"]["monthly_dbus"] == 0
        assert data["total_cost"]["cost_per_month"] == 0

    def test_disabled_dimensions_contribute_zero(self):
        data = calculate_agent_evaluation_cost(
            _request(
                labels_enabled=False,
                input_tokens_millions=999,
                output_tokens_millions=999,
                synthetic_data_enabled=True,
                synthetic_questions=2,
            ),
            db=_PricingDb(),
        )["data"]
        assert list(data["usage_calculation"]) == ["synthetic_questions"]
        assert data["dbu_calculation"]["monthly_dbus"] == 10

    @pytest.mark.parametrize(
        "overrides",
        [
            {"labels_enabled": False, "synthetic_data_enabled": False},
            {"input_tokens_millions": -1},
            {"output_tokens_millions": float("inf")},
            {"output_tokens_millions": float("nan")},
            {"synthetic_questions": -1},
            {"synthetic_questions": 1.5},
        ],
    )
    def test_request_rejects_invalid_values(self, overrides):
        with pytest.raises(ValidationError):
            _request(**overrides)

    def test_discount_support(self):
        discount = DiscountConfig.model_validate({
            "global": {"dbu_discount": 10},
        })
        data = calculate_agent_evaluation_cost(
            _request(discount_config=discount),
            db=_PricingDb(),
        )["data"]
        assert data["discount_summary"][
            "total_discount_percentage"
        ] == pytest.approx(10, abs=0.05)
        assert data["total_cost"]["total_after_discount"] == pytest.approx(
            data["total_cost"]["cost_per_month"] * 0.9,
            abs=0.01,
        )

    def test_missing_exact_regional_price_is_rejected(self):
        with pytest.raises(HTTPException, match="pricing is not available"):
            calculate_agent_evaluation_cost(
                _request(),
                db=_PricingDb(price=None),
            )

    def test_standard_tier_is_rejected(self):
        with pytest.raises(HTTPException, match="Premium or Enterprise"):
            calculate_agent_evaluation_cost(
                _request(tier="STANDARD"),
                db=_PricingDb(),
            )

    def test_response_is_transparent_and_includes_exclusion_note(self):
        data = calculate_agent_evaluation_cost(
            _request(synthetic_data_enabled=False),
            db=_PricingDb(),
        )["data"]
        assert data["workload_type"] == "AGENT_EVALUATION"
        assert data["configuration"]["labels_enabled"] is True
        assert "synthetic_questions" not in data["usage_calculation"]
        assert any(
            "exclude the evaluated app or model inference" in note
            and "Model Serving or Foundation Model API" in note
            for note in data["notes"]
        )


class TestPersistence:
    fields = {
        "agent_evaluation_labels_enabled": True,
        "agent_evaluation_input_tokens_millions": 2.5,
        "agent_evaluation_output_tokens_millions": 1.25,
        "agent_evaluation_synthetic_data_enabled": True,
        "agent_evaluation_synthetic_questions": 50,
    }

    def test_create_folds_public_fields_into_json(self):
        item = LineItemCreate.model_validate({
            "estimate_id": uuid4(),
            "workload_name": "Evaluation",
            "workload_type": "AGENT_EVALUATION",
            **self.fields,
        })
        mapped = map_ai_parse_api_fields(
            item.model_dump(),
            item.model_fields_set,
        )
        validate_agent_evaluation_workload_config(
            mapped["workload_type"],
            mapped["workload_config"],
        )
        assert mapped["workload_config"] == self.fields
        assert not (
            set(AGENT_EVALUATION_CONFIG_FIELDS) & mapped.keys()
        )
        assert not any(
            hasattr(LineItem, field)
            for field in AGENT_EVALUATION_CONFIG_FIELDS
        )
        LineItem(**mapped)

    def test_update_merges_and_response_hydrates(self):
        update = LineItemUpdate(
            agent_evaluation_output_tokens_millions=9.75,
            agent_evaluation_synthetic_data_enabled=False,
        )
        mapped = map_ai_parse_api_fields(
            update.model_dump(exclude_unset=True),
            update.model_fields_set,
            existing_workload_config={
                **self.fields,
                "unrelated": {"nested": True},
            },
        )
        assert mapped["workload_config"][
            "agent_evaluation_output_tokens_millions"
        ] == 9.75
        assert mapped["workload_config"][
            "agent_evaluation_synthetic_data_enabled"
        ] is False
        assert mapped["workload_config"]["unrelated"] == {"nested": True}

        now = datetime.now(timezone.utc)
        response = LineItemResponse.model_validate({
            "line_item_id": uuid4(),
            "estimate_id": uuid4(),
            "workload_name": "Evaluation",
            "workload_type": "AGENT_EVALUATION",
            "workload_config": mapped["workload_config"],
            "created_at": now,
            "updated_at": now,
        })
        assert response.agent_evaluation_labels_enabled is True
        assert response.agent_evaluation_input_tokens_millions == 2.5
        assert response.agent_evaluation_output_tokens_millions == 9.75
        assert response.agent_evaluation_synthetic_data_enabled is False
        assert response.agent_evaluation_synthetic_questions == 50

    def test_switching_type_strips_evaluation_fields(self):
        update = LineItemUpdate(workload_type="JOBS")
        mapped = map_ai_parse_api_fields(
            update.model_dump(exclude_unset=True),
            update.model_fields_set,
            existing_workload_config={
                **self.fields,
                "unrelated": {"nested": True},
            },
        )
        assert mapped["workload_config"] == {
            "unrelated": {"nested": True},
        }

    @pytest.mark.parametrize(
        ("field", "value", "message"),
        [
            (
                "agent_evaluation_labels_enabled",
                "true",
                "boolean",
            ),
            (
                "agent_evaluation_input_tokens_millions",
                -1,
                "greater than",
            ),
            (
                "agent_evaluation_output_tokens_millions",
                float("inf"),
                "finite",
            ),
            (
                "agent_evaluation_synthetic_questions",
                1.5,
                "integer",
            ),
        ],
    )
    def test_json_validation_rejects_invalid_types_and_values(
        self,
        field,
        value,
        message,
    ):
        with pytest.raises(ValueError, match=message):
            validate_agent_evaluation_workload_config(
                "AGENT_EVALUATION",
                {**self.fields, field: value},
            )

    def test_json_validation_requires_enabled_dimension_quantities(self):
        for missing_field in (
            "agent_evaluation_input_tokens_millions",
            "agent_evaluation_output_tokens_millions",
            "agent_evaluation_synthetic_questions",
        ):
            config = dict(self.fields)
            config.pop(missing_field)
            with pytest.raises(ValueError, match="required"):
                validate_agent_evaluation_workload_config(
                    "AGENT_EVALUATION",
                    config,
                )

    def test_json_validation_requires_enabled_feature(self):
        with pytest.raises(ValueError, match="At least one"):
            validate_agent_evaluation_workload_config(
                "AGENT_EVALUATION",
                {
                    **self.fields,
                    "agent_evaluation_labels_enabled": False,
                    "agent_evaluation_synthetic_data_enabled": False,
                },
            )

    @pytest.mark.parametrize("clone_kind", ["estimate", "line_item"])
    def test_clones_deep_copy_workload_config(self, clone_kind):
        original = LineItem(
            estimate_id=uuid4(),
            workload_name="Evaluation",
            workload_type="AGENT_EVALUATION",
            workload_config={
                **self.fields,
                "unrelated": {"nested": ["value"]},
            },
        )
        if clone_kind == "estimate":
            cloned = _copy_line_item(original, uuid4())
        else:
            cloned = _copy_line_item_for_clone(
                original,
                2,
                "Evaluation Copy",
            )
        assert cloned.workload_config == original.workload_config
        assert cloned.workload_config is not original.workload_config
        cloned.workload_config["unrelated"]["nested"].append("changed")
        assert original.workload_config["unrelated"]["nested"] == ["value"]


class TestRegistriesAndAssistant:
    def test_workload_and_sku_registries(self):
        workload = next(
            item
            for item in DEFAULT_WORKLOAD_TYPES
            if item["workload_type"] == "AGENT_EVALUATION"
        )
        assert workload["display_name"] == "Agent Evaluation"
        assert workload["sku_product_type_standard"] == (
            "SERVERLESS_REAL_TIME_INFERENCE"
        )
        assert workload["display_order"] == 16
        assert get_calculation_sku("AGENT_EVALUATION") == (
            "SERVERLESS_REAL_TIME_INFERENCE"
        )
        assert get_service_sku("AGENT_EVALUATION") == (
            "SERVERLESS_REAL_TIME_INFERENCE"
        )
        assert get_export_sku(
            SimpleNamespace(workload_type="AGENT_EVALUATION"),
            "aws",
        ) == "SERVERLESS_REAL_TIME_INFERENCE"
        for relative_path in (
            "scripts/functions/01_Utility_Functions.py",
            "etl/lakebase_setup/functions/01_Utility_Functions.py",
        ):
            source = (ROOT / relative_path).read_text(encoding="utf-8")
            assert "WHEN 'AGENT_EVALUATION' THEN" in source

    def test_assistant_schema_defaults_prompt_and_response_mapping(self):
        tool = next(t for t in TOOLS if t["name"] == "propose_workload")
        properties = tool["parameters"]["properties"]
        tool_fields = {
            field.removeprefix("agent_evaluation_")
            for field in AGENT_EVALUATION_CONFIG_FIELDS
        }
        assert tool_fields.issubset(properties)
        assert "AGENT_EVALUATION" in properties["workload_type"]["enum"]
        assert "evaluated app/model inference is excluded" in SYSTEM_PROMPT

        agent = EstimateAgent(None)
        agent.current_estimate = {
            "cloud": "aws",
            "region": "us-east-1",
            "tier": "PREMIUM",
        }
        result = agent._propose_workload(
            "AGENT_EVALUATION",
            "Evaluation",
            labels_enabled=True,
            input_tokens_millions=2.5,
            output_tokens_millions=1.25,
            synthetic_data_enabled=True,
            synthetic_questions=50,
        )
        proposal = result["proposed_workload"]
        for field, value in self._fields().items():
            assert proposal[f"agent_evaluation_{field}"] == value

        defaults = agent._apply_defaults({
            "workload_type": "AGENT_EVALUATION",
            "cloud": "aws",
        })
        assert defaults["agent_evaluation_labels_enabled"] is True
        assert defaults["agent_evaluation_input_tokens_millions"] == 1
        assert defaults["agent_evaluation_output_tokens_millions"] == 1
        assert defaults["agent_evaluation_synthetic_data_enabled"] is False
        assert defaults["agent_evaluation_synthetic_questions"] == 0

    @staticmethod
    def _fields():
        return {
            "labels_enabled": True,
            "input_tokens_millions": 2.5,
            "output_tokens_millions": 1.25,
            "synthetic_data_enabled": True,
            "synthetic_questions": 50,
        }

    @pytest.mark.parametrize(
        "relative_path",
        [
            "scripts/install_lakemeter.py",
            "scripts/notebooks/02_create_database.py",
        ],
    )
    def test_installer_seed_uses_json_without_schema_columns(
        self,
        relative_path,
    ):
        source = (ROOT / relative_path).read_text(encoding="utf-8")
        assert '"AGENT_EVALUATION"' in source
        for field in AGENT_EVALUATION_CONFIG_FIELDS:
            assert f"{field} VARCHAR" not in source
            assert f'{field}", "NUMERIC' not in source

    def test_data_update_is_idempotent_and_has_no_migration(self):
        source = (
            ROOT
            / "scripts/upgrades/data_updates/022_agent_evaluation.sql"
        ).read_text(encoding="utf-8")
        assert "'AGENT_EVALUATION'" in source
        assert "ON CONFLICT (workload_type) DO UPDATE" in source
        assert "ALTER TABLE" not in source.upper()

    def test_etl_contains_canonical_serverless_component_rates(self):
        source = (
            ROOT
            / "etl/pricing_sync/10_Load_Serverless_Product_Rates.ipynb"
        ).read_text(encoding="utf-8")
        for component, rate in (
            ("input_tokens", "2.143"),
            ("output_tokens", "8.571"),
            ("synthetic_questions", "5.0"),
        ):
            assert f'\\"size_or_model\\": \\"{component}\\"' in source
            assert f'\\"dbu_rate\\": {rate}' in source


class TestExcelExport:
    config = {
        "agent_evaluation_labels_enabled": True,
        "agent_evaluation_input_tokens_millions": 2,
        "agent_evaluation_output_tokens_millions": 3,
        "agent_evaluation_synthetic_data_enabled": True,
        "agent_evaluation_synthetic_questions": 4,
    }

    def test_export_calculation_matches_api(self):
        item = make_item(
            workload_type="AGENT_EVALUATION",
            workload_name="Evaluation",
            workload_config=self.config,
        )
        _, _, _, total_dbus, _ = calc_item_values(
            item,
            False,
            False,
            0,
            "aws",
            [],
        )
        assert total_dbus == pytest.approx(49.999)

    def test_excel_writes_separate_formula_backed_rows(self):
        item = make_item(
            workload_type="AGENT_EVALUATION",
            workload_name="Evaluation",
            workload_config=self.config,
        )
        output = build_estimate_excel(
            make_estimate(),
            [item],
            "aws",
            "ap-southeast-1",
            "PREMIUM",
        )
        payload = BytesIO(output.getvalue())
        formulas = openpyxl.load_workbook(
            payload,
            data_only=False,
        ).active
        payload.seek(0)
        cached = openpyxl.load_workbook(payload, data_only=True).active

        expected = (
            ("Evaluation – Evaluation Input Tokens", 2, 2.143),
            ("Evaluation – Evaluation Output Tokens", 3, 8.571),
            ("Evaluation – Synthetic Data Questions", 4, 5.0),
        )
        rows = []
        for name, quantity, rate in expected:
            row = find_row_by_name(formulas, name)
            assert row is not None
            rows.append(row)
            expected_dbus = quantity * rate
            expected_cost = expected_dbus * SRTI_SINGAPORE_PRICE
            assert formulas.cell(row, COL_TOKEN_QTY).value == quantity
            assert formulas.cell(row, COL_DBU_PER_M).value == rate
            assert formulas.cell(row, COL_DBUS_MO).value == (
                f"=N{row}*O{row}"
            )
            assert cached.cell(row, COL_DBUS_MO).value == pytest.approx(
                expected_dbus
            )
            assert formulas.cell(row, COL_DBU_RATE).value == 0.088
            assert formulas.cell(row, COL_DBU_COST_L).value == (
                f"=Q{row}*R{row}"
            )
            assert cached.cell(row, COL_DBU_COST_L).value == pytest.approx(
                expected_cost
            )
            assert formulas.cell(row, COL_TOTAL_L).value == (
                f"=U{row}+Y{row}+AE{row}"
            )
            assert cached.cell(row, COL_TOTAL_L).value == pytest.approx(
                expected_cost
            )
            config = formulas.cell(row, COL_CONFIG).value
            assert f"Quantity: {quantity:g}" in config
            assert f"Canonical rate: {rate:.3f}" in config
            assert "evaluated app or model inference" in formulas.cell(
                row,
                COL_NOTES,
            ).value

        assert rows == list(range(rows[0], rows[0] + 3))
        assert find_row_by_name(formulas, "Evaluation") is None
        totals_row = find_totals_row(formulas)
        assert formulas.cell(totals_row, COL_DBUS_MO).value == (
            f"=SUM(Q{rows[0]}:Q{rows[-1]})"
        )
        assert formulas.cell(totals_row, COL_TOTAL_L).value == (
            f"=SUM(AF{rows[0]}:AF{rows[-1]})"
        )
        assert sum(
            cached.cell(row, COL_DBUS_MO).value for row in rows
        ) == pytest.approx(49.999)

    @pytest.mark.parametrize(
        (
            "labels_enabled",
            "synthetic_enabled",
            "expected_names",
        ),
        [
            (
                True,
                False,
                [
                    "Evaluation – Evaluation Input Tokens",
                    "Evaluation – Evaluation Output Tokens",
                ],
            ),
            (
                False,
                True,
                ["Evaluation – Synthetic Data Questions"],
            ),
        ],
    )
    def test_excel_one_component_behavior(
        self,
        labels_enabled,
        synthetic_enabled,
        expected_names,
    ):
        config = {
            **self.config,
            "agent_evaluation_labels_enabled": labels_enabled,
            "agent_evaluation_synthetic_data_enabled": synthetic_enabled,
        }
        output = build_estimate_excel(
            make_estimate(),
            [
                make_item(
                    workload_type="AGENT_EVALUATION",
                    workload_name="Evaluation",
                    workload_config=config,
                )
            ],
            "aws",
            "us-east-1",
            "PREMIUM",
        )
        sheet = openpyxl.load_workbook(
            output,
            data_only=False,
        ).active
        all_names = [
            "Evaluation – Evaluation Input Tokens",
            "Evaluation – Evaluation Output Tokens",
            "Evaluation – Synthetic Data Questions",
        ]
        for name in all_names:
            if name in expected_names:
                assert find_row_by_name(sheet, name) is not None
            else:
                assert find_row_by_name(sheet, name) is None
        assert find_row_by_name(sheet, "Evaluation") is None

    def test_excel_labels_include_zero_quantity_rows(self):
        config = {
            **self.config,
            "agent_evaluation_input_tokens_millions": 0,
            "agent_evaluation_output_tokens_millions": 0,
            "agent_evaluation_synthetic_data_enabled": False,
        }
        output = build_estimate_excel(
            make_estimate(),
            [
                make_item(
                    workload_type="AGENT_EVALUATION",
                    workload_name="Zero",
                    workload_config=config,
                )
            ],
            "aws",
            "us-east-1",
            "PREMIUM",
        )
        sheet = openpyxl.load_workbook(output, data_only=True).active
        for name in (
            "Zero – Evaluation Input Tokens",
            "Zero – Evaluation Output Tokens",
        ):
            row = find_row_by_name(sheet, name)
            assert row is not None
            assert sheet.cell(row, COL_DBUS_MO).value == 0

    def test_excel_rejects_missing_exact_region(self):
        item = make_item(
            workload_type="AGENT_EVALUATION",
            workload_name="Evaluation",
            workload_config=self.config,
        )
        with pytest.raises(ValueError, match="pricing is not available"):
            build_estimate_excel(
                make_estimate(),
                [item],
                "aws",
                "unsupported-region",
                "PREMIUM",
            )

    def test_excel_rejects_standard_tier(self):
        item = make_item(
            workload_type="AGENT_EVALUATION",
            workload_name="Evaluation",
            workload_config=self.config,
        )
        with pytest.raises(ValueError, match="Premium or Enterprise"):
            build_estimate_excel(
                make_estimate(),
                [item],
                "aws",
                "us-east-1",
                "STANDARD",
            )


class TestReleaseWiring:
    def test_release_keeps_version_and_includes_agent_evaluation_update(self):
        manifest = json.loads(
            (ROOT / "scripts/upgrades/release.json").read_text(
                encoding="utf-8"
            )
        )
        assert manifest["version"] == "0.2.0"
        action = next(
            item
            for item in manifest["data_updates"]
            if item["id"] == "022-agent-evaluation"
        )
        assert action["path"] == (
            "scripts/upgrades/data_updates/022_agent_evaluation.sql"
        )
        assert len(action["sha256"]) == 64
        assert len(manifest["runtime_sha256"]) == 64
