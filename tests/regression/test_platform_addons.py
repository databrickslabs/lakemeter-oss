from datetime import date
from io import BytesIO
from types import SimpleNamespace

import openpyxl
import pytest

from app.routes.export import build_estimate_excel
from app.routes.calculate.platform_addon_calc import (
    PlatformAddonCalculationRequest,
    calculate_platform_addon,
)
from app.models.estimate import Estimate
from app.schemas.estimate import EstimateCreate, EstimateResponse, EstimateUpdate
from app.services.ai_agent import EstimateAgent
from app.services.platform_addons import (
    calculate_platform_addon_cost,
    get_platform_addon_rate,
    get_selected_platform_addon,
    validate_platform_addon_selection,
)
from tests.export.cross_workload.conftest import make_line_item


@pytest.mark.parametrize(
    ("cloud", "tier", "expected_rate"),
    [
        ("AWS", "ENTERPRISE", 15),
        ("AZURE", "PREMIUM", 10),
        ("GCP", "ENTERPRISE", 15),
    ],
)
def test_enhanced_security_rates_are_cloud_specific(
    cloud,
    tier,
    expected_rate,
):
    result = calculate_platform_addon_cost(
        1_000,
        "ENHANCED_SECURITY_COMPLIANCE",
        cloud,
        tier,
    )
    assert result["applied_rate_pct"] == expected_rate
    assert result["cost"] == expected_rate * 10


def test_mission_critical_uses_active_promotion_then_standard_rate():
    promotional = get_platform_addon_rate(
        "MISSION_CRITICAL",
        "AWS",
        "ENTERPRISE",
        pricing_date=date(2027, 6, 30),
    )
    standard = get_platform_addon_rate(
        "MISSION_CRITICAL",
        "AWS",
        "ENTERPRISE",
        pricing_date=date(2027, 7, 1),
    )

    assert promotional["applied_rate_pct"] == 15
    assert promotional["standard_rate_pct"] == 30
    assert promotional["promotion"]["end_date"] == "2027-06-30"
    assert standard["applied_rate_pct"] == 30
    assert standard["promotion"] is None


def test_addon_discount_applies_after_list_spend_uplift():
    result = calculate_platform_addon_cost(
        1_000,
        "MISSION_CRITICAL",
        "AWS",
        "ENTERPRISE",
        pricing_date=date(2027, 6, 30),
        discount_pct=20,
    )

    assert result["cost_before_discount"] == 150
    assert result["discount_amount"] == 30
    assert result["cost"] == 120


def test_calculation_endpoint_returns_basis_rate_and_cost():
    response = calculate_platform_addon(
        PlatformAddonCalculationRequest(
            cloud="AZURE",
            tier="PREMIUM",
            addon_type="ENHANCED_SECURITY_COMPLIANCE",
            product_spend_at_list=2_000,
        )
    )

    assert response["success"] is True
    assert response["data"]["basis"] == "product_spend_at_list"
    assert response["data"]["applied_rate_pct"] == 10
    assert response["data"]["cost"] == 200


def test_mission_critical_is_unavailable_on_gcp():
    with pytest.raises(ValueError, match="not available on GCP"):
        get_platform_addon_rate(
            "MISSION_CRITICAL",
            "GCP",
            "ENTERPRISE",
        )


def test_selection_rejects_stacked_or_ineligible_addons():
    with pytest.raises(ValueError, match="Select only one"):
        get_selected_platform_addon(
            {
                "platform_addons": [
                    "ENHANCED_SECURITY_COMPLIANCE",
                    "MISSION_CRITICAL",
                ]
            }
        )
    with pytest.raises(ValueError, match="requires ENTERPRISE"):
        validate_platform_addon_selection(
            {"platform_addons": ["ENHANCED_SECURITY_COMPLIANCE"]},
            "AWS",
            "PREMIUM",
        )


def test_estimate_api_schemas_and_model_include_saved_selection():
    config = {"platform_addons": ["ENHANCED_SECURITY_COMPLIANCE"]}
    created = EstimateCreate(
        estimate_name="Regulated workload",
        cloud="AWS",
        tier="ENTERPRISE",
        discount_config=config,
    )
    updated = EstimateUpdate(discount_config=config)

    assert created.discount_config == config
    assert updated.discount_config == config
    assert "discount_config" in Estimate.__table__.columns
    assert "discount_config" in EstimateResponse.model_fields


def test_ai_summary_includes_addon_but_excludes_vm_from_its_basis():
    agent = object.__new__(EstimateAgent)
    agent.current_estimate = {
        "estimate_name": "Mission critical estimate",
        "cloud": "AWS",
        "region": "us-east-1",
        "tier": "ENTERPRISE",
        "discount_config": {"platform_addons": ["MISSION_CRITICAL"]},
    }
    agent.current_workloads = [
        {
            "workload_name": "Workload",
            "workload_type": "JOBS",
            "dbu_cost": 100,
            "dsu_cost": 50,
            "vm_cost": 1_000,
            "total_cost": 1_150,
        }
    ]
    agent.proposed_workloads = []

    addon = agent._get_platform_addon_summary()
    summary = agent._get_estimate_summary()

    assert addon["product_spend_at_list"] == 150
    assert addon["cost"] == 22.5
    assert summary["total_monthly_cost"] == "$1172.50"
    assert summary["estimate"]["platform_addon"]["display_name"] == (
        "Mission Critical"
    )


def test_excel_addon_uses_dbu_dsu_list_spend_and_excludes_vm():
    estimate = SimpleNamespace(
        estimate_name="Platform Add-on",
        status="draft",
        version=1,
        created_at=None,
        updated_at=None,
        discount_config={
            "platform_addons": ["MISSION_CRITICAL"],
            "global": {"platform_addon_discount": 20},
        },
    )
    item = make_line_item(
        workload_type="JOBS",
        workload_name="Classic Jobs",
        driver_node_type="m5.xlarge",
        worker_node_type="m5.xlarge",
        num_workers=2,
        hours_per_month=100,
    )

    payload = build_estimate_excel(
        estimate,
        [item],
        "aws",
        "us-east-1",
        "ENTERPRISE",
    ).getvalue()
    formulas = openpyxl.load_workbook(BytesIO(payload), data_only=False)
    values = openpyxl.load_workbook(BytesIO(payload), data_only=True)
    formula_sheet = formulas["Databricks Estimate"]
    value_sheet = values["Databricks Estimate"]

    def find_label(label):
        return next(
            row
            for row in range(1, formula_sheet.max_row + 1)
            if formula_sheet.cell(row, 1).value == label
        )

    workload_summary_row = find_label(
        "WORKLOAD COST SUMMARY (BEFORE PLATFORM ADD-ONS)"
    )
    addon_section_row = find_label("PLATFORM ADD-ON")
    final_summary_row = find_label("FINAL ESTIMATE SUMMARY")
    totals_row = find_label("WORKLOAD TOTALS:")
    product_spend_row = find_label("Product Spend at List")
    addon_list_row = find_label("Add-on Cost (List)")
    addon_discounted_row = find_label("Platform Add-on Cost")
    final_estimate_row = find_label("FINAL ESTIMATE")

    def numeric(row, column):
        value = value_sheet.cell(row, column).value
        return value if isinstance(value, (int, float)) else 0

    product_spend = sum(
        numeric(row, 21) + numeric(row, 25)
        for row in range(1, workload_summary_row)
    )
    vm_cost = sum(
        numeric(row, 31)
        for row in range(1, workload_summary_row)
    )
    addon_list_cost = value_sheet.cell(addon_list_row, 2).value
    addon_discounted_cost = value_sheet.cell(
        addon_discounted_row,
        2,
    ).value
    product_formula = formula_sheet.cell(product_spend_row, 2).value
    discounted_formula = formula_sheet.cell(
        addon_discounted_row,
        2,
    ).value

    assert workload_summary_row < addon_section_row < final_summary_row
    assert not any(
        formula_sheet.cell(row, 3).value == "Platform Add-on"
        for row in range(1, workload_summary_row)
    )
    assert vm_cost > 0
    assert addon_list_cost == pytest.approx(product_spend * 0.15)
    assert addon_discounted_cost == pytest.approx(product_spend * 0.15 * 0.8)
    assert addon_list_cost != pytest.approx((product_spend + vm_cost) * 0.15)
    assert product_formula == f"=U{totals_row}+Y{totals_row}"
    assert "AE" not in product_formula
    assert discounted_formula.startswith(f"=B{addon_list_row}*(1-B")
    assert formula_sheet.cell(final_estimate_row, 2).value.startswith("=B")


def test_excel_without_addon_still_shows_zero_addon_and_final_summary():
    estimate = SimpleNamespace(
        estimate_name="No Add-on",
        status="draft",
        version=1,
        created_at=None,
        updated_at=None,
        discount_config=None,
    )
    item = make_line_item(
        workload_type="JOBS",
        workload_name="Classic Jobs",
        driver_node_type="m5.xlarge",
        worker_node_type="m5.xlarge",
        num_workers=2,
        hours_per_month=100,
    )
    payload = build_estimate_excel(
        estimate,
        [item],
        "aws",
        "us-east-1",
        "PREMIUM",
    ).getvalue()
    formulas = openpyxl.load_workbook(
        BytesIO(payload),
        data_only=False,
    )["Databricks Estimate"]
    values = openpyxl.load_workbook(
        BytesIO(payload),
        data_only=True,
    )["Databricks Estimate"]

    selected_row = next(
        row
        for row in range(1, formulas.max_row + 1)
        if formulas.cell(row, 1).value == "Selected Add-on"
    )
    addon_cost_row = next(
        row
        for row in range(1, formulas.max_row + 1)
        if formulas.cell(row, 1).value == "Platform Add-on Cost"
    )
    product_spend_row = next(
        row
        for row in range(1, formulas.max_row + 1)
        if formulas.cell(row, 1).value == "Product Spend at List"
    )

    assert formulas.cell(selected_row, 2).value == "None selected"
    assert values.cell(product_spend_row, 2).value > 0
    assert values.cell(addon_cost_row, 2).value == 0
    assert any(
        formulas.cell(row, 1).value == "FINAL ESTIMATE"
        for row in range(1, formulas.max_row + 1)
    )
