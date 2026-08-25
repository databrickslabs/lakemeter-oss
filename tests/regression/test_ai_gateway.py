"""Regression coverage for Unity AI Gateway pricing and persistence."""
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import openpyxl
import pytest
from pydantic import ValidationError

from app.models.line_item import LineItem
from app.routes.calculate import ai_gateway_calc
from app.routes.calculate.ai_gateway_calc import (
    calculate_ai_gateway_cost,
    calculate_ai_gateway_usage,
)
from app.routes.calculate.helpers import get_sku_type as get_calculation_sku
from app.routes.calculate.schemas import (
    AIGatewayCalculationRequest,
    DiscountConfig,
)
from app.routes.estimates import _copy_line_item
from app.routes.export.excel_builder import build_estimate_excel
from app.routes.export.excel_item_helpers import calc_item_values
from app.routes.export.pricing import _get_sku_type as get_export_sku
from app.routes.line_items import _copy_line_item_for_clone
from app.schemas.line_item import (
    AI_GATEWAY_CONFIG_FIELDS,
    LineItemCreate,
    LineItemResponse,
    LineItemUpdate,
    map_ai_parse_api_fields,
    validate_ai_gateway_workload_config,
)
from app.services.ai_agent import EstimateAgent, TOOLS
from app.services.lakebase_queries import get_sku_type as get_service_sku
from tests.regression.conftest import make_item
from tests.regression.excel_helpers import (
    COL_CONFIG,
    COL_DBU_COST_L,
    COL_DBU_RATE,
    COL_DBUS_MO,
    COL_NOTES,
    COL_TOTAL_L,
    find_dbu_summary_row,
    find_row_by_name,
    find_totals_row,
    make_estimate,
)


ROOT = Path(__file__).resolve().parents[2]
RTI_PRICE = 0.07


class _Result:
    def __init__(self, row):
        self.row = row

    def fetchone(self):
        return self.row


class _PricingDb:
    def __init__(self, price=RTI_PRICE):
        self.price = price

    def execute(self, statement, params):
        if "pt" in params:
            row = (
                SimpleNamespace(price_per_dbu=self.price)
                if self.price is not None
                else None
            )
            return _Result(row)
        if "sku" in params:
            return _Result(("dbu", True))
        raise AssertionError(f"Unexpected query params: {params}")


@pytest.fixture(autouse=True)
def _patch_reference_validation(monkeypatch):
    monkeypatch.setattr(
        ai_gateway_calc, "validate_cloud", lambda *args, **kwargs: None
    )
    monkeypatch.setattr(
        ai_gateway_calc, "validate_region", lambda *args, **kwargs: None
    )
    monkeypatch.setattr(
        ai_gateway_calc, "validate_tier", lambda *args, **kwargs: None
    )


def _request(**kwargs):
    values = {
        "cloud": "AWS",
        "region": "us-east-1",
        "tier": "PREMIUM",
        "inference_tables_enabled": True,
        "inference_tables_input_method": "requests",
        "inference_tables_requests_millions": 2,
        "inference_tables_avg_request_payload_kb": 0.25,
        "inference_tables_avg_response_payload_kb": 0.75,
        "inference_tables_monthly_payload_gb": None,
        "usage_tracking_enabled": True,
        "usage_tracking_input_method": "payload_gb",
        "usage_tracking_requests_millions": None,
        "usage_tracking_avg_request_payload_kb": None,
        "usage_tracking_avg_response_payload_kb": None,
        "usage_tracking_monthly_payload_gb": 5,
    }
    values.update(kwargs)
    return AIGatewayCalculationRequest(**values)


def _usage(**kwargs):
    values = {
        key: value
        for key, value in _request().model_dump().items()
        if key not in {"cloud", "region", "tier", "discount_config"}
    }
    values.update(kwargs)
    return calculate_ai_gateway_usage(**values)


class TestCalculation:
    def test_components_use_independent_payload_volumes(self):
        result = calculate_ai_gateway_cost(_request(), db=_PricingDb())
        data = result["data"]

        inference_usage = data["usage_calculation"]["inference_tables"]
        tracking_usage = data["usage_calculation"]["usage_tracking"]
        assert inference_usage["input_method"] == "requests"
        assert inference_usage["monthly_payload_gb"] == 2.0
        assert inference_usage["monthly_dbus"] == pytest.approx(2.858)
        assert tracking_usage["input_method"] == "payload_gb"
        assert tracking_usage["monthly_payload_gb"] == 5.0
        assert tracking_usage["monthly_dbus"] == pytest.approx(7.145)
        assert len(data["component_breakdown"]) == 2
        assert {
            component["dbu_per_gb"]
            for component in data["component_breakdown"]
        } == {1.429}
        expected_dbus = 10.003
        assert data["dbu_calculation"]["monthly_dbus"] == pytest.approx(
            expected_dbus
        )
        assert data["dbu_calculation"]["dbu_price"] == RTI_PRICE
        assert data["dbu_calculation"]["monthly_dbu_cost"] == pytest.approx(
            expected_dbus * RTI_PRICE
        )
        assert data["sku_type"] == "SERVERLESS_REAL_TIME_INFERENCE"
        assert any(
            "Direct monthly payload GB is preferred" in note
            for note in data["notes"]
        )
        assert any(
            "guardrail evaluator costs are excluded" in note
            for note in data["notes"]
        )

    def test_direct_payload_mode_matches_requests_mode(self):
        request_data = calculate_ai_gateway_cost(
            _request(),
            db=_PricingDb(),
        )["data"]
        direct_data = calculate_ai_gateway_cost(
            _request(
                inference_tables_input_method="payload_gb",
                inference_tables_requests_millions=None,
                inference_tables_avg_request_payload_kb=None,
                inference_tables_avg_response_payload_kb=None,
                inference_tables_monthly_payload_gb=2,
            ),
            db=_PricingDb(),
        )["data"]

        assert direct_data["usage_calculation"]["inference_tables"][
            "monthly_payload_gb"
        ] == 2.0
        assert direct_data["dbu_calculation"]["monthly_dbus"] == pytest.approx(
            request_data["dbu_calculation"]["monthly_dbus"]
        )
        assert direct_data["total_cost"]["cost_per_month"] == pytest.approx(
            request_data["total_cost"]["cost_per_month"]
        )

    @pytest.mark.parametrize(
        ("inference_tables", "usage_tracking", "expected_rate"),
        [
            (True, False, 1.429),
            (False, True, 1.429),
            (True, True, 2.858),
        ],
    )
    def test_feature_toggles(
        self, inference_tables, usage_tracking, expected_rate
    ):
        usage = _usage(
            inference_tables_enabled=inference_tables,
            usage_tracking_enabled=usage_tracking,
            inference_tables_requests_millions=1,
            inference_tables_avg_request_payload_kb=1,
            inference_tables_avg_response_payload_kb=0,
            usage_tracking_input_method="requests",
            usage_tracking_requests_millions=1,
            usage_tracking_avg_request_payload_kb=1,
            usage_tracking_avg_response_payload_kb=0,
            usage_tracking_monthly_payload_gb=None,
        )
        assert usage["monthly_dbus"] == pytest.approx(expected_rate)

    def test_discount_helpers_apply_to_gateway_sku(self):
        discount = DiscountConfig.model_validate({
            "global": {"dbu_discount": 10},
        })
        data = calculate_ai_gateway_cost(
            _request(discount_config=discount),
            db=_PricingDb(),
        )["data"]
        assert data["discount_summary"][
            "total_discount_percentage"
        ] == pytest.approx(10, abs=0.05)
        assert data["total_cost"]["total_after_discount"] == pytest.approx(
            data["total_cost"]["cost_per_month"] * 0.9,
            abs=0.01,
        )

    def test_missing_exact_regional_price_is_rejected(self):
        from fastapi import HTTPException

        with pytest.raises(HTTPException, match="pricing is not available"):
            calculate_ai_gateway_cost(_request(), db=_PricingDb(price=None))

    def test_standard_tier_is_rejected(self):
        from fastapi import HTTPException

        with pytest.raises(HTTPException, match="Premium or Enterprise"):
            calculate_ai_gateway_cost(
                _request(tier="STANDARD"),
                db=_PricingDb(),
            )

    @pytest.mark.parametrize(
        "overrides",
        [
            {
                "inference_tables_enabled": False,
                "usage_tracking_enabled": False,
            },
            {"inference_tables_requests_millions": -1},
            {"inference_tables_avg_request_payload_kb": float("inf")},
            {"usage_tracking_monthly_payload_gb": float("nan")},
            {"usage_tracking_monthly_payload_gb": -1},
            {
                "inference_tables_input_method": "requests",
                "inference_tables_requests_millions": None,
            },
            {
                "usage_tracking_input_method": "payload_gb",
                "usage_tracking_monthly_payload_gb": None,
            },
        ],
    )
    def test_request_rejects_invalid_values(self, overrides):
        with pytest.raises(ValidationError):
            _request(**overrides)

    def test_disabled_component_inputs_may_be_absent(self):
        request = _request(
            usage_tracking_enabled=False,
            usage_tracking_input_method=None,
            usage_tracking_requests_millions=None,
            usage_tracking_avg_request_payload_kb=None,
            usage_tracking_avg_response_payload_kb=None,
            usage_tracking_monthly_payload_gb=None,
        )
        data = calculate_ai_gateway_cost(request, db=_PricingDb())["data"]
        assert [component["component"] for component in data[
            "component_breakdown"
        ]] == ["inference_tables"]


class TestPersistence:
    fields = {
        "ai_gateway_inference_tables_enabled": True,
        "ai_gateway_inference_tables_input_method": "requests",
        "ai_gateway_inference_tables_requests_millions": 0.5,
        "ai_gateway_inference_tables_avg_request_payload_kb": 1.2,
        "ai_gateway_inference_tables_avg_response_payload_kb": 3.4,
        "ai_gateway_inference_tables_monthly_payload_gb": 2.3,
        "ai_gateway_usage_tracking_enabled": False,
        "ai_gateway_usage_tracking_input_method": "requests",
        "ai_gateway_usage_tracking_requests_millions": 0.25,
        "ai_gateway_usage_tracking_avg_request_payload_kb": 0.4,
        "ai_gateway_usage_tracking_avg_response_payload_kb": 0.6,
        "ai_gateway_usage_tracking_monthly_payload_gb": 0.25,
    }

    def test_create_maps_public_fields_only_to_json(self):
        item = LineItemCreate.model_validate({
            "estimate_id": uuid4(),
            "workload_name": "Gateway",
            "workload_type": "AI_GATEWAY",
            **self.fields,
        })
        mapped = map_ai_parse_api_fields(
            item.model_dump(),
            item.model_fields_set,
        )
        validate_ai_gateway_workload_config(
            mapped["workload_type"],
            mapped["workload_config"],
        )

        assert mapped["workload_config"] == self.fields
        assert not (set(AI_GATEWAY_CONFIG_FIELDS) & mapped.keys())
        assert not any(
            hasattr(LineItem, field) for field in AI_GATEWAY_CONFIG_FIELDS
        )
        LineItem(**mapped)

    def test_update_merges_and_response_hydrates(self):
        update = LineItemUpdate(
            ai_gateway_usage_tracking_input_method="payload_gb",
            ai_gateway_usage_tracking_monthly_payload_gb=9.75,
            ai_gateway_usage_tracking_enabled=True,
        )
        mapped = map_ai_parse_api_fields(
            update.model_dump(exclude_unset=True),
            update.model_fields_set,
            existing_workload_config={
                **self.fields,
                "unrelated": {"nested": True},
            },
        )
        assert mapped["workload_config"][
            "ai_gateway_usage_tracking_input_method"
        ] == (
            "payload_gb"
        )
        assert mapped["workload_config"][
            "ai_gateway_usage_tracking_monthly_payload_gb"
        ] == (
            9.75
        )
        assert mapped["workload_config"]["ai_gateway_usage_tracking_enabled"]
        assert mapped["workload_config"]["unrelated"] == {"nested": True}

        now = datetime.now(timezone.utc)
        response = LineItemResponse.model_validate({
            "line_item_id": uuid4(),
            "estimate_id": uuid4(),
            "workload_name": "Gateway",
            "workload_type": "AI_GATEWAY",
            "workload_config": mapped["workload_config"],
            "created_at": now,
            "updated_at": now,
        })
        assert response.ai_gateway_usage_tracking_input_method == "payload_gb"
        assert response.ai_gateway_usage_tracking_monthly_payload_gb == 9.75
        assert response.ai_gateway_inference_tables_requests_millions == 0.5
        assert response.ai_gateway_usage_tracking_enabled is True

    def test_switching_type_strips_gateway_config(self):
        update = LineItemUpdate(workload_type="JOBS")
        mapped = map_ai_parse_api_fields(
            update.model_dump(exclude_unset=True),
            update.model_fields_set,
            existing_workload_config={
                **self.fields,
                "unrelated": {"nested": True},
            },
        )
        assert mapped["workload_config"] == {
            "unrelated": {"nested": True},
        }

    @pytest.mark.parametrize(
        ("field", "value", "message"),
        [
            (
                "ai_gateway_inference_tables_requests_millions",
                -1,
                "greater than",
            ),
            (
                "ai_gateway_inference_tables_avg_request_payload_kb",
                float("inf"),
                "finite",
            ),
            (
                "ai_gateway_usage_tracking_avg_response_payload_kb",
                float("nan"),
                "finite",
            ),
            (
                "ai_gateway_usage_tracking_monthly_payload_gb",
                -1,
                "greater than",
            ),
        ],
    )
    def test_json_validation_rejects_invalid_numbers(
        self, field, value, message
    ):
        with pytest.raises(ValueError, match=message):
            validate_ai_gateway_workload_config(
                "AI_GATEWAY",
                {**self.fields, field: value},
            )

    @pytest.mark.parametrize(
        "config",
        [
            {
                **fields,
                "ai_gateway_inference_tables_input_method": "requests",
                "ai_gateway_inference_tables_requests_millions": None,
            },
            {
                **fields,
                "ai_gateway_usage_tracking_enabled": True,
                "ai_gateway_usage_tracking_input_method": "payload_gb",
                "ai_gateway_usage_tracking_monthly_payload_gb": None,
            },
        ],
    )
    def test_json_validation_requires_mode_specific_values(self, config):
        with pytest.raises(ValueError, match="required"):
            validate_ai_gateway_workload_config("AI_GATEWAY", config)

    def test_json_validation_requires_paid_feature(self):
        with pytest.raises(ValueError, match="At least one"):
            validate_ai_gateway_workload_config(
                "AI_GATEWAY",
                {
                    **self.fields,
                    "ai_gateway_inference_tables_enabled": False,
                    "ai_gateway_usage_tracking_enabled": False,
                },
            )

    def test_json_validation_ignores_absent_disabled_component_inputs(self):
        config = {
            key: value
            for key, value in self.fields.items()
            if not key.startswith("ai_gateway_usage_tracking_")
        }
        config["ai_gateway_usage_tracking_enabled"] = False
        validate_ai_gateway_workload_config("AI_GATEWAY", config)

    @pytest.mark.parametrize("clone_kind", ["estimate", "line_item"])
    def test_clones_deep_copy_workload_config(self, clone_kind):
        original = LineItem(
            estimate_id=uuid4(),
            workload_name="Gateway",
            workload_type="AI_GATEWAY",
            workload_config={
                **self.fields,
                "unrelated": {"nested": ["value"]},
            },
        )
        if clone_kind == "estimate":
            cloned = _copy_line_item(original, uuid4())
        else:
            cloned = _copy_line_item_for_clone(original, 2, "Gateway Copy")

        assert cloned.workload_config == original.workload_config
        assert cloned.workload_config is not original.workload_config
        cloned.workload_config["unrelated"]["nested"].append("changed")
        assert original.workload_config["unrelated"]["nested"] == ["value"]


class TestRegistriesAndAssistant:
    def test_python_and_sql_sku_resolvers(self):
        assert get_calculation_sku("AI_GATEWAY") == (
            "SERVERLESS_REAL_TIME_INFERENCE"
        )
        assert get_service_sku("AI_GATEWAY") == (
            "SERVERLESS_REAL_TIME_INFERENCE"
        )
        assert get_export_sku(
            SimpleNamespace(workload_type="AI_GATEWAY"),
            "aws",
        ) == "SERVERLESS_REAL_TIME_INFERENCE"
        for relative_path in (
            "scripts/functions/01_Utility_Functions.py",
            "etl/lakebase_setup/functions/01_Utility_Functions.py",
        ):
            source = (ROOT / relative_path).read_text(encoding="utf-8")
            assert "WHEN 'AI_GATEWAY' THEN" in source

    def test_assistant_schema_and_proposal_preserve_gateway_fields(self):
        tool = next(t for t in TOOLS if t["name"] == "propose_workload")
        properties = tool["parameters"]["properties"]
        tool_fields = {
            field.removeprefix("ai_gateway_")
            for field in AI_GATEWAY_CONFIG_FIELDS
        }
        assert tool_fields.issubset(properties)
        assert "AI_GATEWAY" in properties["workload_type"]["enum"]

        agent = EstimateAgent(None)
        agent.current_estimate = {
            "cloud": "aws",
            "region": "us-east-1",
            "tier": "PREMIUM",
        }
        result = agent._propose_workload(
            "AI_GATEWAY",
            "Gateway",
            **self._assistant_fields(),
        )
        proposal = result["proposed_workload"]
        for field, value in self._assistant_fields().items():
            assert proposal[f"ai_gateway_{field}"] == value

    @staticmethod
    def _assistant_fields():
        return {
            "inference_tables_enabled": False,
            "inference_tables_input_method": "requests",
            "inference_tables_requests_millions": 0.5,
            "inference_tables_avg_request_payload_kb": 1.2,
            "inference_tables_avg_response_payload_kb": 3.4,
            "inference_tables_monthly_payload_gb": 2.3,
            "usage_tracking_enabled": True,
            "usage_tracking_input_method": "payload_gb",
            "usage_tracking_requests_millions": 0.5,
            "usage_tracking_avg_request_payload_kb": 1.2,
            "usage_tracking_avg_response_payload_kb": 3.4,
            "usage_tracking_monthly_payload_gb": 11.25,
        }


class TestExportAndInstaller:
    config = {
        "ai_gateway_inference_tables_enabled": True,
        "ai_gateway_inference_tables_input_method": "payload_gb",
        "ai_gateway_inference_tables_requests_millions": 2,
        "ai_gateway_inference_tables_avg_request_payload_kb": 0.25,
        "ai_gateway_inference_tables_avg_response_payload_kb": 0.75,
        "ai_gateway_inference_tables_monthly_payload_gb": 2,
        "ai_gateway_usage_tracking_enabled": True,
        "ai_gateway_usage_tracking_input_method": "payload_gb",
        "ai_gateway_usage_tracking_requests_millions": 5,
        "ai_gateway_usage_tracking_avg_request_payload_kb": 0.25,
        "ai_gateway_usage_tracking_avg_response_payload_kb": 0.75,
        "ai_gateway_usage_tracking_monthly_payload_gb": 5,
    }

    @pytest.mark.parametrize(
        "config",
        [
            config,
            {
                **config,
                "ai_gateway_inference_tables_input_method": "requests",
                "ai_gateway_usage_tracking_input_method": "requests",
            },
        ],
    )
    def test_export_calculation_matches_api_for_both_modes(self, config):
        item = make_item(
            workload_type="AI_GATEWAY",
            workload_name="Gateway",
            workload_config=config,
        )
        _, _, _, total_dbus, _ = calc_item_values(
            item, False, False, 0, "aws", []
        )
        assert total_dbus == pytest.approx(10.003)

    def test_excel_writes_component_rows_and_safe_totals(self):
        item = make_item(
            workload_type="AI_GATEWAY",
            workload_name="Gateway",
            workload_config=self.config,
        )
        output = build_estimate_excel(
            make_estimate(),
            [item],
            "aws",
            "us-east-1",
            "PREMIUM",
        )
        payload = BytesIO(output.getvalue())
        formulas = openpyxl.load_workbook(payload, data_only=False).active
        payload.seek(0)
        cached = openpyxl.load_workbook(payload, data_only=True).active
        inference_row = find_row_by_name(
            formulas,
            "Gateway – Inference Tables",
        )
        usage_row = find_row_by_name(
            formulas,
            "Gateway – Usage Tracking",
        )

        assert inference_row is not None
        assert usage_row == inference_row + 1
        assert find_row_by_name(formulas, "Gateway") is None
        for row, payload_gb in (
            (inference_row, 2),
            (usage_row, 5),
        ):
            expected_dbus = payload_gb * 1.429
            expected_cost = expected_dbus * RTI_PRICE
            assert formulas.cell(row, COL_DBUS_MO).value == pytest.approx(
                expected_dbus
            )
            assert cached.cell(row, COL_DBUS_MO).value == pytest.approx(
                expected_dbus
            )
            assert formulas.cell(row, COL_DBU_RATE).value == RTI_PRICE
            assert formulas.cell(row, COL_DBU_COST_L).value == (
                f"=Q{row}*R{row}"
            )
            assert cached.cell(row, COL_DBU_COST_L).value == pytest.approx(
                expected_cost
            )
            assert formulas.cell(row, COL_TOTAL_L).value == (
                f"=U{row}+AA{row}"
            )
            config = formulas.cell(row, COL_CONFIG).value
            assert f"Monthly payload: {payload_gb} GB" in config
            assert "Component rate: 1.429 DBU/GB" in config
            notes = formulas.cell(row, COL_NOTES).value
            assert "Direct monthly payload GB is preferred" in notes
            assert "guardrail evaluator costs are excluded" in notes

        totals_row = find_totals_row(formulas)
        dbu_summary_row = find_dbu_summary_row(formulas)
        assert formulas.cell(totals_row, COL_DBUS_MO).value == (
            f"=SUM(Q{inference_row}:Q{usage_row})"
        )
        assert sum(
            cached.cell(row, COL_DBUS_MO).value
            for row in (inference_row, usage_row)
        ) == pytest.approx(10.003)
        assert formulas.cell(totals_row, COL_TOTAL_L).value == (
            f"=SUM(AB{inference_row}:AB{usage_row})"
        )
        assert formulas.cell(dbu_summary_row, 3).value == (
            f"=SUM(Q{inference_row}:Q{usage_row})"
        )

    @pytest.mark.parametrize(
        ("inference_tables", "usage_tracking", "expected_name"),
        [
            (True, False, "Gateway – Inference Tables"),
            (False, True, "Gateway – Usage Tracking"),
        ],
    )
    def test_excel_writes_only_enabled_component(
        self,
        inference_tables,
        usage_tracking,
        expected_name,
    ):
        config = {
            **self.config,
            "ai_gateway_inference_tables_enabled": inference_tables,
            "ai_gateway_usage_tracking_enabled": usage_tracking,
        }
        output = build_estimate_excel(
            make_estimate(),
            [
                make_item(
                    workload_type="AI_GATEWAY",
                    workload_name="Gateway",
                    workload_config=config,
                )
            ],
            "aws",
            "us-east-1",
            "PREMIUM",
        )
        sheet = openpyxl.load_workbook(output, data_only=False).active
        assert find_row_by_name(sheet, expected_name) is not None
        unexpected_name = (
            "Gateway – Usage Tracking"
            if inference_tables
            else "Gateway – Inference Tables"
        )
        assert find_row_by_name(sheet, unexpected_name) is None
        assert find_row_by_name(sheet, "Gateway") is None

    def test_excel_rejects_missing_exact_region(self):
        item = make_item(
            workload_type="AI_GATEWAY",
            workload_name="Gateway",
            workload_config=self.config,
        )
        with pytest.raises(ValueError, match="pricing is not available"):
            build_estimate_excel(
                make_estimate(),
                [item],
                "aws",
                "unsupported-region",
                "PREMIUM",
            )

    @pytest.mark.parametrize(
        "relative_path",
        [
            "scripts/install_lakemeter.py",
            "scripts/notebooks/02_create_database.py",
        ],
    )
    def test_installer_seed_has_no_gateway_schema_columns(
        self, relative_path
    ):
        source = (ROOT / relative_path).read_text(encoding="utf-8")
        assert '"AI_GATEWAY"' in source
        for field in AI_GATEWAY_CONFIG_FIELDS:
            assert f"{field} VARCHAR" not in source
            assert f'{field}", "NUMERIC' not in source

    def test_data_update_is_idempotent_and_has_no_migration(self):
        source = (
            ROOT
            / "scripts/upgrades/data_updates/021_ai_gateway.sql"
        ).read_text(encoding="utf-8")
        assert "'AI_GATEWAY'" in source
        assert "ON CONFLICT (workload_type) DO UPDATE" in source
        assert "ALTER TABLE" not in source.upper()
